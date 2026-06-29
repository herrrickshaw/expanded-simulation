import pandas as pd
import pickle
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

with open("monte_carlo_summary.pkl", "rb") as f:
    summary = pickle.load(f)
draws_df = pd.read_csv("monte_carlo_draws.csv")

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
BIG_NUM = Font(name=FONT_NAME, bold=True, size=20, color="1F4E78")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CARD_FILL = PatternFill("solid", start_color="EAF1F8")
HUNG_FILL = PatternFill("solid", start_color="F8CBAD")
MAJORITY_FILL = PatternFill("solid", start_color="C6E0B4")

wb = load_workbook("coalition_model.xlsx")
ws = wb.create_sheet("Monte Carlo Simulation")
ws.sheet_view.showGridLines = False

n_sims = summary["n_simulations"]
nda_maj = summary["nda_majority_count"]
india_maj = summary["india_majority_count"]
hung = summary["hung_count"]

ws.cell(1, 1, f"Monte Carlo Simulation: {n_sims:,} Randomized Seat Distributions").font = TITLE_FONT
ws.merge_cells("A1:H1")
ws.cell(2, 1, "Each draw randomizes every party's seats within a historically plausible range (e.g. NDA 220-360, YSRCP 0-25), constrained to sum to 543. Existing parties only.").font = NOTE_FONT
ws.merge_cells("A2:H2")

# KPI cards
row = 4
kpis = [
    ("NDA Outright Majority", nda_maj/n_sims, MAJORITY_FILL),
    ("INDIA Outright Majority", india_maj/n_sims, MAJORITY_FILL),
    ("Hung Parliament", hung/n_sims, HUNG_FILL),
]
col = 1
for label, pct, fill in kpis:
    ws.cell(row, col, label).font = NORMAL
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    c = ws.cell(row+1, col, pct)
    c.font = BIG_NUM
    c.number_format = "0.0%"
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+1)
    for rr in range(row, row+3):
        for cc in range(col, col+2):
            ws.cell(rr, cc).fill = fill
            ws.cell(rr, cc).border = BORDER
    col += 3

row2 = row + 4
ws.cell(row2, 1, f"Based on {n_sims:,} simulations with random seed fixed for reproducibility (seed=42).").font = NOTE_FONT
ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=8)

# ---- Histogram of NDA seats ----
hrow = row2 + 2
ws.cell(hrow, 1, "Distribution of NDA Seats Across All Simulations").font = SUBTITLE_FONT
hrow += 1

bins = list(range(200, 360, 10))
draws_df["nda_bin"] = pd.cut(draws_df["NDA"], bins=bins, right=True)
hist = draws_df["nda_bin"].value_counts().sort_index()

headers = ["NDA Seat Range", "Count", "% of Simulations"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(hrow, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER
hrow += 1
hist_start = hrow
for interval, count in hist.items():
    label = f"{int(interval.left)+1}-{int(interval.right)}"
    ws.cell(hrow, 1, label).font = NORMAL
    ws.cell(hrow, 2, int(count)).font = NORMAL
    ws.cell(hrow, 3, count/n_sims).font = NORMAL
    ws.cell(hrow, 3).number_format = "0.0%"
    for c in range(1, 4):
        ws.cell(hrow, c).border = BORDER
    # highlight the >=272 majority threshold bin
    if int(interval.left) < 272 <= int(interval.right):
        for c in range(1, 4):
            ws.cell(hrow, c).fill = PatternFill("solid", start_color="FFF2CC")
    hrow += 1
hist_end = hrow - 1

ws.cell(hrow, 1, "Majority threshold (272 seats) falls within the highlighted band.").font = NOTE_FONT
ws.merge_cells(start_row=hrow, start_column=1, end_row=hrow, end_column=3)

chart = BarChart()
chart.type = "col"
chart.title = "Distribution of NDA Seats (20,000 simulations)"
chart.y_axis.title = "Number of simulations"
chart.x_axis.title = "NDA seat range"
data_ref = Reference(ws, min_col=2, min_row=hist_start-1, max_row=hist_end)
cats_ref = Reference(ws, min_col=1, min_row=hist_start, max_row=hist_end)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 9
chart.width = 18
ws.add_chart(chart, "E4")

# ---- Average Banzhaf power conditional on hung parliament ----
brow = hist_end + 3
ws.cell(brow, 1, "Average Banzhaf Power Index, Conditional on Hung Parliament").font = SUBTITLE_FONT
brow += 1
ws.cell(brow, 1, "(Averaged only across the ~38% of simulations that produced a hung parliament; shows realistic bargaining power IF no bloc wins outright.)").font = NOTE_FONT
brow += 1

headers2 = ["Party", "Avg Banzhaf Index (when hung)", "# Hung Draws Appeared In"]
for c, h in enumerate(headers2, start=1):
    cell = ws.cell(brow, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER
brow += 1
avg_banzhaf = summary["avg_banzhaf"]
banzhaf_acc = summary["banzhaf_accumulator"]
for p, v in sorted(avg_banzhaf.items(), key=lambda x: -x[1]):
    ws.cell(brow, 1, p).font = NORMAL
    ws.cell(brow, 2, round(v, 4)).font = NORMAL
    ws.cell(brow, 2).number_format = "0.00%"
    ws.cell(brow, 3, len(banzhaf_acc[p])).font = NORMAL
    for c in range(1, 4):
        ws.cell(brow, c).border = BORDER
    brow += 1

col_widths = [22, 14, 16, 4, 24, 16, 16, 16]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("coalition_model.xlsx")
print("Monte Carlo sheet done.")
