import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
NORMAL = Font(name=FONT_NAME, size=10)
BOLD = Font(name=FONT_NAME, bold=True, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = load_workbook("election_ml_tcpd_real.xlsx")
df = pd.read_csv("ae_national_bloc_trend.csv")

ws = wb.create_sheet("Long-Term AE Trend (1961-2023)", 1)
ws.sheet_view.showGridLines = False
ws.cell(1, 1, "62-Year State Assembly Bloc Trend, Real Data (1961-2023)").font = TITLE_FONT
ws.merge_cells("A1:G1")
ws.cell(2, 1, "56,358 real State Assembly seat-winner rows from TCPD AE dataset, aggregated to national NDA/INDIA/Other bloc share per year. Pre-1980 NDA figures include Jana Sangh, BJP's direct predecessor.").font = NOTE_FONT
ws.merge_cells("A2:G2")

headers = ["Year","NDA Seats","INDIA Seats","Other Seats","Total Seats","NDA Share","INDIA Share"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(4, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

start_row = 5
for i, row in df.iterrows():
    r = start_row + i
    ws.cell(r, 1, int(row["Year"])).font = NORMAL
    ws.cell(r, 2, int(row["NDA"])).font = NORMAL
    ws.cell(r, 3, int(row["INDIA"])).font = NORMAL
    ws.cell(r, 4, int(row["Other"])).font = NORMAL
    ws.cell(r, 5, int(row["Total"])).font = NORMAL
    ws.cell(r, 6, row["NDA_Share"]).font = NORMAL
    ws.cell(r, 6).number_format = "0.0%"
    ws.cell(r, 7, row["INDIA_Share"]).font = NORMAL
    ws.cell(r, 7).number_format = "0.0%"
    for c in range(1, 8):
        ws.cell(r, c).border = BORDER
last_row = start_row + len(df) - 1

col_widths = [8, 12, 12, 12, 12, 12, 12]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

chart = LineChart()
chart.title = "NDA vs INDIA Bloc Share of State Assembly Seats, 1961-2023 (real data)"
chart.y_axis.title = "Share of seats won"
chart.x_axis.title = "Year"
data_ref = Reference(ws, min_col=6, max_col=7, min_row=4, max_row=last_row)
cats_ref = Reference(ws, min_col=1, min_row=start_row, max_row=last_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 11
chart.width = 26
ws.add_chart(chart, "I4")

wb.save("election_ml_tcpd_real.xlsx")
print(f"AE trend sheet done. rows {start_row}-{last_row}")
