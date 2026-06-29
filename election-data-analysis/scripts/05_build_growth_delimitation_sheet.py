import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
GREEN_LINK = Font(name=FONT_NAME, size=10, color="008000")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RED_FILL = PatternFill("solid", start_color="F8CBAD")
GREEN_FILL = PatternFill("solid", start_color="C6E0B4")

wb = load_workbook("election_analysis.xlsx")
df = pd.read_csv("state_data_raw.csv")
df_states = df[df["Assembly_Seats"].notna()].copy()  # only states with assemblies (excludes small UTs)
df_states = df_states.sort_values("LS_Seats_2024", ascending=False).reset_index(drop=True)

ws_src = wb["State-wise LS Data"]

ws3 = wb.create_sheet("Growth & Delimitation")
ws3.sheet_view.showGridLines = False
ws3.cell(1, 1, "Population Growth vs. Representation: Which States Are Over/Under-Represented").font = TITLE_FONT
ws3.merge_cells("A1:G1")
ws3.cell(2, 1, "Basis: LS seats frozen since 1976 at 1971-Census proportions; current constituency boundaries based on 2001 Census; population shown is 2011 Census (latest available).").font = NOTE_FONT
ws3.merge_cells("A2:G2")

headers = ["State", "LS Seats", "2011 Census Population", "Pop per Seat (2011)",
           "All-India Avg Pop/Seat (2011)", "Index vs All-India Avg (=100)", "Representation Status"]
for c, h in enumerate(headers, start=1):
    cell = ws3.cell(4, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

start_row = 5
n = len(df_states)
last_row = start_row + n - 1

# We need the all-India average pop/seat as an absolute reference computed across STATES only (excl. small UTs without assembly)
total_pop = df_states["Census2011_Pop"].sum()
total_seats = df_states["LS_Seats_2024"].sum()

for i, row in df_states.iterrows():
    r = start_row + i
    ws3.cell(r, 1, row["State"]).font = NORMAL
    ws3.cell(r, 2, int(row["LS_Seats_2024"])).font = NORMAL
    ws3.cell(r, 3, int(row["Census2011_Pop"])).font = NORMAL
    ws3.cell(r, 3).number_format = "#,##0"
    ws3.cell(r, 4, f"=C{r}/B{r}").font = NORMAL
    ws3.cell(r, 4).number_format = "#,##0"
    ws3.cell(r, 5, f"=SUM($C${start_row}:$C${last_row})/SUM($B${start_row}:$B${last_row})").font = NORMAL
    ws3.cell(r, 5).number_format = "#,##0"
    ws3.cell(r, 6, f"=D{r}/E{r}*100").font = NORMAL
    ws3.cell(r, 6).number_format = "0.0"
    ws3.cell(r, 7, f'=IF(F{r}>110,"Under-represented (large seat population)",IF(F{r}<90,"Over-represented (small seat population)","Broadly proportionate"))').font = NORMAL
    for c in range(1, 8):
        ws3.cell(r, c).border = BORDER

from openpyxl.formatting.rule import CellIsRule
ws3.conditional_formatting.add(f"F{start_row}:F{last_row}",
    CellIsRule(operator="greaterThan", formula=["110"], fill=RED_FILL))
ws3.conditional_formatting.add(f"F{start_row}:F{last_row}",
    CellIsRule(operator="lessThan", formula=["90"], fill=GREEN_FILL))

legend_r = last_row + 2
ws3.cell(legend_r, 1, "Index > 110 (red): state's avg constituency carries notably MORE people per seat than national average → under-represented").font = NOTE_FONT
ws3.cell(legend_r+1, 1, "Index < 90 (green): state's avg constituency carries notably FEWER people per seat → over-represented relative to its population").font = NOTE_FONT
ws3.merge_cells(start_row=legend_r, start_column=1, end_row=legend_r, end_column=7)
ws3.merge_cells(start_row=legend_r+1, start_column=1, end_row=legend_r+1, end_column=7)

col_widths = [22, 11, 20, 16, 22, 20, 32]
for i, w in enumerate(col_widths, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = "A5"

# Chart: bar chart of Index ranked
chart = BarChart()
chart.type = "bar"
chart.title = "Representation Index by State (2011 Census pop per LS seat, All-India = 100)"
chart.y_axis.title = "State"
chart.x_axis.title = "Index"
data_ref = Reference(ws3, min_col=6, min_row=4, max_row=last_row)
cats_ref = Reference(ws3, min_col=1, min_row=start_row, max_row=last_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 16
chart.width = 18
ws3.add_chart(chart, "I4")

wb.save("election_analysis.xlsx")
print("Sheet 4 (Growth & Delimitation) done. rows:", start_row, last_row)
