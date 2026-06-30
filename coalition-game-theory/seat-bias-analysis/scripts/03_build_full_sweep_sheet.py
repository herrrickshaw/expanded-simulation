import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
NORMAL = Font(name=FONT_NAME, size=10)
BOLD = Font(name=FONT_NAME, bold=True, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HIGHLIGHT_FILL = PatternFill("solid", start_color="FFF2CC")

wb = load_workbook("seat_bias_analysis.xlsx")
df = pd.read_csv("/home/claude/coalition_model/seat_bias_sweep_v2_to1000.csv")

ws = wb.create_sheet("Full Sweep Data & Charts")
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Full Sweep: 534-1,000 Seats, Every 10 Seats Plus Key Endpoints").font = TITLE_FONT
ws.merge_cells("A1:I1")
ws.cell(2, 1, "1,500 simulations per seat-total, 49 seat-totals tested. NDA treated as the incumbent/leading bloc throughout. Rows for 850 (actual proposed ceiling) and 1,000 (stress test) highlighted.").font = NOTE_FONT
ws.merge_cells("A2:I2")

headers = ["Total Seats","Quota","NDA Seat Share","NDA Vote Share","Mean Bonus (pp)",
           "Bonus Volatility (std)","Gallagher Index","P(NDA Majority)"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(4, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

start_row = 5
for i, row in df.iterrows():
    r = start_row + i
    is_highlight = int(row["total_seats"]) in (850, 1000)
    ws.cell(r, 1, int(row["total_seats"])).font = BOLD if is_highlight else NORMAL
    ws.cell(r, 2, int(row["quota"])).font = NORMAL
    ws.cell(r, 3, row["nda_seat_share"]).font = NORMAL
    ws.cell(r, 3).number_format = "0.00%"
    ws.cell(r, 4, row["nda_vote_share"]).font = NORMAL
    ws.cell(r, 4).number_format = "0.00%"
    ws.cell(r, 5, row["seat_bonus_leader"]*100).font = NORMAL
    ws.cell(r, 5).number_format = "0.00"
    ws.cell(r, 6, row["bonus_std_across_elections"]*100).font = BOLD
    ws.cell(r, 6).number_format = "0.00"
    ws.cell(r, 7, row["gallagher_index"]).font = NORMAL
    ws.cell(r, 7).number_format = "0.00"
    ws.cell(r, 8, row["pct_nda_majority"]).font = BOLD if is_highlight else NORMAL
    ws.cell(r, 8).number_format = "0.0%"
    for c in range(1, 9):
        ws.cell(r, c).border = BORDER
        if is_highlight:
            ws.cell(r, c).fill = HIGHLIGHT_FILL
last_row = start_row + len(df) - 1

col_widths = [11, 9, 14, 14, 14, 16, 14, 14]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

chart1 = LineChart()
chart1.title = "Bonus Volatility (std across elections) vs. Total Seats, 534-1,000"
chart1.y_axis.title = "Bonus std (percentage points)"
chart1.x_axis.title = "Total seats"
data_ref = Reference(ws, min_col=6, min_row=4, max_row=last_row)
cats_ref = Reference(ws, min_col=1, min_row=start_row, max_row=last_row)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.height = 9
chart1.width = 18
ws.add_chart(chart1, f"A{last_row+3}")

chart2 = LineChart()
chart2.title = "Mean Winner's Bonus vs. Total Seats (flat throughout, 534-1,000)"
chart2.y_axis.title = "Mean bonus (percentage points)"
chart2.x_axis.title = "Total seats"
data_ref2 = Reference(ws, min_col=5, min_row=4, max_row=last_row)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.height = 9
chart2.width = 18
ws.add_chart(chart2, f"K{last_row+3}")

chart3 = LineChart()
chart3.title = "P(NDA Majority) vs Total Seats, 534-1,000 -- RISES steadily, no reversal"
chart3.y_axis.title = "P(NDA majority)"
chart3.x_axis.title = "Total seats"
data_ref3 = Reference(ws, min_col=8, min_row=4, max_row=last_row)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref)
chart3.height = 9
chart3.width = 18
ws.add_chart(chart3, f"A{last_row+22}")

wb.save("seat_bias_analysis.xlsx")
print(f"Full sweep sheet rebuilt. rows {start_row}-{last_row}")
