import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

with open("scenario_results.pkl", "rb") as f:
    results = pickle.load(f)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, color="404040")
BOLD = Font(name=FONT_NAME, bold=True, size=11)
NORMAL = Font(name=FONT_NAME, size=10)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CARD_FILL = PatternFill("solid", start_color="EAF1F8")
HUNG_FILL = PatternFill("solid", start_color="F8CBAD")
MAJORITY_FILL = PatternFill("solid", start_color="C6E0B4")

wb = load_workbook("coalition_model.xlsx")
ws = wb.create_sheet("Dashboard", 0)
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "India Lok Sabha: Coalition Game Theory Dashboard").font = TITLE_FONT
ws.merge_cells("A1:H1")
ws.cell(2, 1, "Existing-parties-only scenario simulation | Majority threshold: 272 of 543 seats | Compiled June 2026").font = SUBTITLE_FONT
ws.merge_cells("A2:H2")

scenario_order = ["S0_Actual_Current","S1_2024_AsDeclared","S2_NDA_Falls_Short",
                  "S3_Deep_Hung_Parliament","S4_NDA_Needs_Swing","S5_INDIA_Largest_No_Majority"]
scenario_labels = {
    "S0_Actual_Current": "S0: Actual Current",
    "S1_2024_AsDeclared": "S1: 2024 Declared",
    "S2_NDA_Falls_Short": "S2: NDA Short (Hung)",
    "S3_Deep_Hung_Parliament": "S3: Deep Fragmentation",
    "S4_NDA_Needs_Swing": "S4: NDA + 1 Swing Party",
    "S5_INDIA_Largest_No_Majority": "S5: INDIA Largest (Hung)",
}

row = 4
ws.cell(row, 1, "Scenario Outcomes at a Glance").font = BOLD
row += 1
headers = ["Scenario", "NDA Seats", "INDIA Seats", "Largest Single Bloc", "Outcome", "# Min. Winning Coalitions"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER
row += 1
start_row = row
for s in scenario_order:
    pw = dict(results[s]["players"])
    nda = pw.get("NDA", 0)
    india = pw.get("INDIA", 0)
    largest = "NDA" if nda >= india else "INDIA"
    is_hung = results[s]["is_hung"]
    outcome = "HUNG PARLIAMENT" if is_hung else f"{largest} has outright majority"
    ws.cell(row, 1, scenario_labels[s]).font = NORMAL
    ws.cell(row, 2, nda).font = NORMAL
    ws.cell(row, 3, india).font = NORMAL
    ws.cell(row, 4, largest).font = NORMAL
    c5 = ws.cell(row, 5, outcome)
    c5.font = BOLD
    c5.fill = HUNG_FILL if is_hung else MAJORITY_FILL
    ws.cell(row, 6, len(results[s]["minimal_winning_coalitions"])).font = NORMAL
    for c in range(1, 7):
        ws.cell(row, c).border = BORDER
    row += 1
last_row = row - 1

row += 2
ws.cell(row, 1, "Key Findings").font = BOLD
row += 1
findings = [
    "1. TODAY'S ACTUAL POSITION (S0): NDA holds 318 of 543 seats on its own — a standalone majority with",
    "   46-seat headroom. No coalition bargaining is structurally necessary; NDA's Banzhaf power = 1.0 and",
    "   every other party's = 0, because NDA alone already satisfies the winning condition.",
    "",
    "2. The moment NDA drops below 272 alone (S2-S5), the game changes completely. Regional/unaligned",
    "   parties that currently have ZERO formal power suddenly become decisive. In S2 (NDA at 255), YSRCP's",
    "   Banzhaf index jumps to 0.193 — meaningful leverage from just 10 seats — because it sits at the",
    "   fulcrum between two near-tied blocs.",
    "",
    "3. Minimal winning coalitions multiply fast under fragmentation: scenario S0 has exactly 1 (NDA alone);",
    "   S5 (INDIA largest but short) has 100. More viable coalition combinations generally means LESS",
    "   stable governance — more potential veto points, more re-bargaining risk over an MP's term.",
    "",
    "4. Small parties are not equally powerful just because they're all 'swing'. In every hung scenario,",
    "   the largest unaligned party (YSRCP-scale, 10-25 seats) captures most of the bargaining power among",
    "   minor parties; 1-seat independents are almost never individually pivotal — they only matter as part",
    "   of an aggregated bloc.",
    "",
    "5. Bottom line on your question: a hung parliament is NOT a likely outcome from current numbers (NDA's",
    "   46-seat majority cushion is large by historical standards), but it is NOT structurally impossible —",
    "   it requires the same kind of seat erosion that already happened to BJP once (303 -> 240 from 2019",
    "   to 2024). The game-theoretic lesson is less 'will it happen' and more 'if it does, who gains power':",
    "   regional and unaligned parties — not new entrants — are the most likely beneficiaries.",
]
for line in findings:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

row += 1
ws.cell(row, 1, "See 'Scenarios' for full seat tables, 'Minimal Winning Coalitions' for all viable governing combinations,").font = NOTE_FONT
row += 1
ws.cell(row, 1, "and 'Banzhaf Power Index' for bargaining-power comparison across scenarios.").font = NOTE_FONT

col_widths = [22, 12, 12, 16, 24, 18, 4, 4]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("coalition_model.xlsx")
print("Dashboard done. Sheet order:", wb.sheetnames)
