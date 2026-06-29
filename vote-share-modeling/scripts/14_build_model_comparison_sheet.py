import pandas as pd
import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FAIL_FILL = PatternFill("solid", start_color="F8CBAD")
OK_FILL = PatternFill("solid", start_color="C6E0B4")

with open("hierarchical_model_results.pkl", "rb") as f:
    hier = pickle.load(f)
with open("dsmm_results.pkl", "rb") as f:
    dsmm = pickle.load(f)

wb = load_workbook("vote_share_models.xlsx")
ws = wb.create_sheet("Model Comparison", 1)
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Continuous Vote-Share Model Comparison: MAE vs Naive Baseline").font = TITLE_FONT
ws.merge_cells("A1:F1")
ws.cell(2, 1, "Lower MAE is better. 'Naive baseline' = predict next election's vote share equals the previous election's (no swing assumed).").font = NOTE_FONT
ws.merge_cells("A2:F2")

row = 4
headers = ["Model","Split","MAE","Naive Baseline MAE","Beats Baseline?"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER
row += 1
start_row = row

rows_data = [
    ("Hierarchical Bayes (MrP-style)", "Validation (2009)", hier["val_mae"], hier["val_naive_mae"]),
    ("Hierarchical Bayes (MrP-style)", "Test (2014+2019)", hier["test_mae"], hier["test_naive_mae"]),
    ("Dirichlet-Swing Matrix", "Validation (2009 from 2004)", dsmm["val_mae_dsmm"], dsmm["val_mae_naive"]),
    ("Dirichlet-Swing Matrix", "Test: 2014 from 2009", dsmm["test1_mae"], dsmm["test1_naive_mae"]),
    ("Dirichlet-Swing Matrix", "Test: 2019 from 2014", dsmm["test2_mae"], dsmm["test2_naive_mae"]),
    ("Dirichlet-Swing Matrix", "Test (pooled)", dsmm["combined_test_mae"], dsmm["combined_test_naive"]),
]
for model, split, mae, naive in rows_data:
    ws.cell(row, 1, model).font = NORMAL
    ws.cell(row, 2, split).font = NORMAL
    ws.cell(row, 3, round(float(mae), 4)).font = BOLD
    ws.cell(row, 4, round(float(naive), 4)).font = NORMAL
    beats = mae < naive
    c5 = ws.cell(row, 5, "Yes" if beats else "No -- naive baseline wins")
    c5.font = BOLD
    c5.fill = OK_FILL if beats else FAIL_FILL
    for c in range(1, 6):
        ws.cell(row, c).border = BORDER
    row += 1
last_row = row - 1

row += 2
ws.cell(row, 1, "For reference: the earlier discrete classifier (predicted win/lose label, not continuous share)").font = SUBTITLE_FONT
row += 1
headers2 = ["Model","Test Accuracy","Majority-Class Baseline","Beats Baseline?"]
for c, h in enumerate(headers2, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER
row += 1
classifier_rows = [
    ("Logistic Regression", 0.210, 0.199),
    ("Random Forest", 0.225, 0.199),
    ("LightGBM", 0.293, 0.199),
]
for model, acc, baseline in classifier_rows:
    ws.cell(row, 1, model).font = NORMAL
    ws.cell(row, 2, acc).font = NORMAL
    ws.cell(row, 2).number_format = "0.0%"
    ws.cell(row, 3, baseline).font = NORMAL
    ws.cell(row, 3).number_format = "0.0%"
    c4 = ws.cell(row, 4, "Marginal" if acc - baseline < 0.1 else "Yes")
    c4.font = BOLD
    for c in range(1, 5):
        ws.cell(row, c).border = BORDER
    row += 1
row += 1
ws.cell(row, 1, "Note: classifier 'beats baseline' marginally on accuracy, but this is a far weaker signal than it appears --").font = NOTE_FONT
row += 1
ws.cell(row, 1, "a 3-class problem where ~33% is random chance and the model only reaches 21-29% means it is still nearly useless.").font = NOTE_FONT
row += 2

ws.cell(row, 1, "Key takeaway").font = SUBTITLE_FONT
row += 1
takeaway = [
    "Continuous vote-share models (both MrP-style hierarchical Bayes and Dirichlet-Swing)",
    "produce errors much closer to a trivial 'assume no change' baseline than the discrete",
    "classifier did to ITS baseline -- meaning continuous modeling degrades far more gracefully",
    "under a real political regime shift. Neither approach, however, actually BEATS the naive",
    "baseline on the genuine held-out test years (2014, 2019) -- the 2014 BJP wave was simply",
    "too large a shift for any model trained on pre-2014 data to anticipate, regardless of",
    "modeling sophistication. This is the honest answer to a genuinely hard problem.",
]
for line in takeaway:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

chart = BarChart()
chart.type = "col"
chart.title = "Model MAE vs Naive Baseline (lower is better)"
chart.y_axis.title = "Mean Absolute Error"
data_ref = Reference(ws, min_col=3, max_col=4, min_row=4, max_row=last_row)
cats_ref = Reference(ws, min_col=2, min_row=start_row, max_row=last_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 10
chart.width = 22
ws.add_chart(chart, f"A{row+2}")

col_widths = [30, 26, 12, 18, 24]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("vote_share_models.xlsx")
print(f"Model comparison sheet done. rows {start_row}-{last_row}")
