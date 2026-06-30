from openpyxl import load_workbook
from openpyxl.styles import Font

FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
RED_BOLD = Font(name=FONT_NAME, bold=True, size=10, color="C00000")

wb = load_workbook("seat_bias_analysis.xlsx")
ws = wb.create_sheet("Methodology", 1)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 112

r = 1
ws.cell(r, 1, "Methodology").font = TITLE_FONT
r += 2

notes = [
    ("Question being answered", BOLD),
    ("  'Is there a seat-total between 534 and 850 where electoral outcomes become more", NORMAL),
    ("  unbiased toward the existing majority in parliament?' Interpreted using the", NORMAL),
    ("  standard political-science 'winner's bonus' measure: the leading bloc's seat share", NORMAL),
    ("  minus its true vote share. NDA (today's actual incumbent/leading bloc) is used", NORMAL),
    ("  throughout as the bloc being checked for bias.", NORMAL),
    ("", NORMAL),
    ("A real bug found and fixed mid-analysis (documented for transparency)", RED_BOLD),
    ("  The FIRST version of this simulation measured 'vote share' using the SAME random", NORMAL),
    ("  draw that determined each seat's winner -- meaning seat share and vote share", NORMAL),
    ("  tracked each other almost perfectly by construction, showing a near-zero bonus", NORMAL),
    ("  (~0.0001) at every seat-total. This is NOT how a real winner's bonus arises and the", NORMAL),
    ("  result was discarded. FIX: rebuilt the simulation so each constituency draws an", NORMAL),
    ("  actual continuous vote-share vector (Dirichlet distribution centered on the state's", NORMAL),
    ("  2024 bloc shares); the WINNER is whichever bloc has the highest drawn share in that", NORMAL),
    ("  seat (the real FPTP mechanism), while the NATIONAL vote share is the seat-weighted", NORMAL),
    ("  average of the underlying continuous shares, independent of who won each seat. This", NORMAL),
    ("  is what allows seat share and vote share to genuinely diverge, producing a real,", NORMAL),
    ("  non-trivial winner's bonus (~1.1-1.2 percentage points) consistent with published", NORMAL),
    ("  estimates in the disproportionality literature for SMP/FPTP systems.", NORMAL),
    ("", NORMAL),
    ("Simulation design", BOLD),
    ("  33 seat-totals tested (534, 544, 554, ... 844, 850 -- every 10 seats plus the exact", NORMAL),
    ("  endpoint). 1,500 simulations per seat-total. Each simulated seat is assigned to a", NORMAL),
    ("  state by that state's CURRENT seat-count weight (today's actual geographic", NORMAL),
    ("  footprint, not a delimitation-style reallocation -- this question is about House", NORMAL),
    ("  SIZE holding geography fixed, a different question from the earlier delimitation", NORMAL),
    ("  seat-redistribution work). Within each seat, a Dirichlet(alpha = state_mean_share *", NORMAL),
    ("  18) draw produces the (NDA, INDIA, Other) vote-share vector for that constituency;", NORMAL),
    ("  concentration=18 was chosen to produce realistic constituency-level variance.", NORMAL),
    ("", NORMAL),
    ("Two distinct metrics, deliberately kept separate", BOLD),
    ("  MEAN BONUS: averaged across all 1,500 simulated elections at a given seat-total --", NORMAL),
    ("  answers 'on average, how biased is this seat-total.'", NORMAL),
    ("  BONUS VOLATILITY (std): how much the bonus varies FROM election to election at a", NORMAL),
    ("  given seat-total -- answers 'how likely is an unusually distorted result in any ONE", NORMAL),
    ("  election.' These can move independently, and in this analysis they do: mean bonus is", NORMAL),
    ("  flat across seat-totals while volatility falls sharply with more seats.", NORMAL),
    ("", NORMAL),
    ("Limitations", BOLD),
    ("  - This holds India's GEOGRAPHIC seat distribution fixed at today's pattern; it does", NORMAL),
    ("    not model the Delimitation Bill's state-by-state reallocation (see the separate", NORMAL),
    ("    'coalition-game-theory' project folder for that question).", NORMAL),
    ("  - Uses only NDA/INDIA/Other as the three competing blocs, consistent with prior work", NORMAL),
    ("    in this project; a finer party breakdown could show different micro-patterns.", NORMAL),
    ("  - The Dirichlet concentration parameter (18) is a modeling choice calibrated to", NORMAL),
    ("    produce plausible constituency-level variance, not estimated from raw vote totals", NORMAL),
    ("    (which were not available at this aggregation level).", NORMAL),
]
for text, font in notes:
    ws.cell(r, 1, text).font = font
    r += 1

wb.save("seat_bias_analysis.xlsx")
print("Methodology sheet done.")
