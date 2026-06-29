import pandas as pd
import pickle
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
RED_TITLE = Font(name=FONT_NAME, bold=True, size=12, color="C00000")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)

wb = Workbook()
ws0 = wb.active
ws0.title = "README"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 112

r = 1
ws0.cell(r, 1, "Continuous Vote-Share Modeling: MrP-Style Hierarchical Bayes + Dirichlet-Swing").font = TITLE_FONT
r += 2

notes = [
    ("What changed from the previous (discrete classifier) approach", BOLD),
    ("  The prior attempt predicted a discrete WIN/LOSE label (NDA/INDIA/Other) per constituency", NORMAL),
    ("  and failed (test accuracy near or below baseline) because hard labels flip entirely and", NORMAL),
    ("  unpredictably across a political regime change. This version predicts the CONTINUOUS", NORMAL),
    ("  vote share V_{p,c} for each bloc p in each constituency c -- the actual quantity", NORMAL),
    ("  requested -- using two methodologically distinct approaches from the literature.", NORMAL),
    ("", NORMAL),
    ("Approach 1: Hierarchical Bayesian model (MrP-inspired partial pooling)", BOLD),
    ("  Adapts the core logic of Cerina & Duch (2021, PLOS ONE) -- multilevel regression with", NORMAL),
    ("  partial pooling across geography -- to our actual data. We do NOT have individual-level", NORMAL),
    ("  survey microdata or census stratification frames as in that paper (which used IHDS,", NORMAL),
    ("  Census, INES, and online convenience samples); this is a genuine, stated scope reduction.", NORMAL),
    ("  Model: V_NDA[c] ~ Normal(alpha_national + alpha_state[s(c)] + beta*lag_V_NDA[c], sigma),", NORMAL),
    ("  with alpha_state partially pooled toward the national mean (the core MrP 'borrowing", NORMAL),
    ("  strength' mechanism). Fit with PyMC (NUTS sampler).", NORMAL),
    ("", NORMAL),
    ("Approach 2: Dirichlet-Swing Matrix Model (Mitra 2026, PLOS ONE)", BOLD),
    ("  Implements the paper's better-performing swing model: a Dirichlet-distributed", NORMAL),
    ("  transition matrix describing where each bloc's previous-election plurality", NORMAL),
    ("  redistributes in the current election. Two real bugs were found and fixed while", NORMAL),
    ("  implementing this -- see 'Dirichlet-Swing Diagnostics' sheet for full detail.", NORMAL),
    ("", NORMAL),
    ("HEADLINE RESULT: neither approach beats a naive 'no swing' baseline on the genuine", RED_TITLE),
    ("chronological test set (2014, 2019) -- a real and important finding, not hidden.", RED_TITLE),
    ("", NORMAL),
    ("  Hierarchical Bayes: test MAE 0.135 vs naive baseline 0.120 (naive wins)", NORMAL),
    ("  Dirichlet-Swing: test MAE 0.155 (pooled) vs naive baseline 0.126 (naive wins)", NORMAL),
    ("  Both, however, are MUCH closer to baseline than the discrete classifier's near-random", NORMAL),
    ("  performance -- continuous modeling degrades gracefully under regime change; discrete", NORMAL),
    ("  classification breaks completely. That is itself a genuine, useful methodological", NORMAL),
    ("  finding from this exercise.", NORMAL),
    ("", NORMAL),
    ("Why: the same root cause as before, now visible mechanistically", BOLD),
    ("  Mean NDA vote share shifted from 0.331 (1999-2004, train era) to 0.407 (2014-2019, test", NORMAL),
    ("  era) -- a genuine +7.6 percentage point national-level swing the models have zero prior", NORMAL),
    ("  information about. The hierarchical model's 90% credible intervals only achieved 13.5%", NORMAL),
    ("  ACTUAL coverage on test -- badly overconfident, because its uncertainty is calibrated", NORMAL),
    ("  to in-regime noise, not to the possibility of a national-level level-shift.", NORMAL),
    ("", NORMAL),
    ("What this means for predicting a new (post-delimitation) configuration", BOLD),
    ("  All three modeling families tried in this project (discrete classification, Bayesian", NORMAL),
    ("  hierarchical regression, Dirichlet-Swing) point to the same conclusion: constituency-", NORMAL),
    ("  level structural/historical features cannot reliably anticipate a genuine political", NORMAL),
    ("  realignment, regardless of how sophisticated the statistical machinery is. Any", NORMAL),
    ("  'prediction' of a new configuration should be read as a structural projection under a", NORMAL),
    ("  stated assumption (e.g. 'if current alignment persists'), not a forecast.", NORMAL),
]
for text, font in notes:
    ws0.cell(r, 1, text).font = font
    r += 1

wb.save("vote_share_models.xlsx")
print("README done.")
