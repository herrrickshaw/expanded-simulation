import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
RED_BOLD = Font(name=FONT_NAME, bold=True, size=11, color="C00000")
BIG_NUM = Font(name=FONT_NAME, bold=True, size=18, color="1F4E78")
DISCLAIMER_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
DISCLAIMER_FONT = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CARD_FILL = PatternFill("solid", start_color="EAF1F8")
RED_FILL = PatternFill("solid", start_color="F8CBAD")
DISCLAIMER_FILL = PatternFill("solid", start_color="C00000")

df = pd.read_csv("/home/claude/coalition_model/seat_bias_sweep_v2_to1000.csv")

wb = load_workbook("seat_bias_analysis.xlsx")
ws = wb.create_sheet("Direct Answer", 0)
ws.sheet_view.showGridLines = False

row = 1
ws.cell(row, 1, "** IMPORTANT DISCLAIMER -- READ BEFORE THE RESULTS BELOW **").font = DISCLAIMER_TITLE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
ws.cell(row, 1).fill = DISCLAIMER_FILL
ws.cell(row, 1).alignment = Alignment(horizontal="center")
row += 1

disclaimer_lines = [
    "The 'volatility falls as seats increase' finding on this sheet is REAL, but it does",
    "NOT mean a bigger House is more fair. This needs to be stated as bluntly as possible:",
    "",
    "If every seat were a FAIR 50/50 coin flip, then whether you flip the coin 534 times",
    "or 1,000 times, the PROBABILITY of any individual flip landing on a given side is",
    "EXACTLY 50% either way -- flipping more times does not change that probability even",
    "slightly. More flips only narrows the SPREAD of the total count around the expected",
    "50/50 split. It does NOT shift the odds toward fairness, because the odds were never",
    "biased by the flip COUNT in the first place -- only by whether the coin itself is fair.",
    "",
    "The same logic applies here. India's seats are NOT fair 50/50 coins: NDA's underlying",
    "support sits at roughly 53%, not 50%. That means MORE seats does not pull the system",
    "toward 50/50 fairness -- it pulls the result MORE TIGHTLY toward NDA's existing",
    "advantage, making an NDA majority MORE CERTAIN, not less. The data below CONFIRMS",
    "this directly, now tested all the way to 1,000 seats: NDA's probability of keeping",
    "its majority RISES from 96.9% (534 seats) to 99.5% (1,000 seats) -- correlation",
    "+0.918 over the full range. More seats does not level the playing field here; it",
    "raises the floor under whoever is already ahead, and the effect does not reverse or",
    "plateau anywhere in the range tested.",
    "",
    "The 'volatility' finding is therefore better read as: more seats make the system more",
    "PREDICTABLE around whatever mean already exists -- biased or not -- not more FAIR.",
]
for line in disclaimer_lines:
    ws.cell(row, 1, line).font = DISCLAIMER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row, 1).fill = DISCLAIMER_FILL
    ws.cell(row, 1).alignment = Alignment(wrap_text=True)
    row += 1

row += 1

ws.cell(row, 1, "Extended to 1,000 Seats: Does the Pattern Reverse or Find a Sweet Spot?").font = TITLE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1
ws.cell(row, 1, "Answer: No reversal, no plateau. The entrenchment effect continues smoothly past the actual proposed ceiling of 850, all the way to 1,000.").font = NOTE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 2

nda_bonus_mean = df["seat_bonus_leader"].mean()
p_maj_534 = df[df.total_seats==534]["pct_nda_majority"].values[0]
p_maj_850 = df[df.total_seats==850]["pct_nda_majority"].values[0]
p_maj_1000 = df[df.total_seats==1000]["pct_nda_majority"].values[0]
corr_p_maj = np.corrcoef(df["total_seats"], df["pct_nda_majority"])[0,1]
bonus_std_534 = df[df.total_seats==534]["bonus_std_across_elections"].values[0]
bonus_std_1000 = df[df.total_seats==1000]["bonus_std_across_elections"].values[0]
pct_reduction = (1 - bonus_std_1000/bonus_std_534) * 100

kpi_row = row
kpis = [
    ("Avg winner's bonus,\nALL seat-totals 534-1000", f"{nda_bonus_mean*100:.2f}pp\n(flat)", CARD_FILL),
    ("P(NDA majority)\nat 534 seats", f"{p_maj_534*100:.1f}%", CARD_FILL),
    ("P(NDA majority)\nat 850 seats", f"{p_maj_850*100:.1f}%", RED_FILL),
    ("P(NDA majority)\nat 1,000 seats", f"{p_maj_1000*100:.1f}%", RED_FILL),
]
col = 1
for label, val, fill in kpis:
    ws.cell(kpi_row, col, label).font = NORMAL
    ws.cell(kpi_row, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row+1, end_column=col+1)
    c = ws.cell(kpi_row+2, col, val)
    c.font = BIG_NUM
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=kpi_row+2, start_column=col, end_row=kpi_row+3, end_column=col+1)
    for rr in range(kpi_row, kpi_row+4):
        for cc in (col, col+1):
            ws.cell(rr, cc).fill = fill
            ws.cell(rr, cc).border = BORDER
    col += 2
row = kpi_row + 5

ws.cell(row, 1, "Full-range correlations (534-1,000 seats, 49 points tested)").font = SUBTITLE_FONT
row += 1
corr_lines = [
    f"Correlation(seats, P(NDA majority)) = +{corr_p_maj:.3f}  -- entrenchment strengthens",
    "slightly when extended past 850 (was +0.884 over 534-850 alone; +0.918 over 534-1,000).",
    "Correlation(seats, bonus volatility) = -0.966 -- still falling smoothly, now a",
    f"{pct_reduction:.0f}% total reduction in volatility from 534 to 1,000 (was 21.6% over 534-850 alone).",
    "Correlation(seats, mean bonus SIZE) remains ~0 -- the average bonus itself never moves,",
    "staying within 1.11-1.21 percentage points across all 49 seat-totals tested.",
]
for line in corr_lines:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

row += 1
ws.cell(row, 1, "Why 1,000 specifically doesn't change the qualitative answer").font = SUBTITLE_FONT
row += 1
why_lines = [
    "1,000 seats is roughly 1.85x the current 543 -- well beyond even the government's",
    "actual proposed ceiling (850, from the defeated Delimitation Bill, 2026). Testing it",
    "confirms this is NOT a phenomenon specific to the 534-850 'realistic' range -- it's a",
    "structural property of averaging more independent draws around an off-center mean.",
    "There's no seat count where the curve bends back toward 50/50 fairness, because",
    "nothing in the simulation's mechanics would cause that -- the mean (NDA's true ~53%",
    "support) doesn't change with seat count, only the SPREAD around that mean shrinks.",
    "A genuinely toss-up race (50/50 true support) would behave differently -- see the",
    "caveat on this in the Methodology sheet.",
]
for line in why_lines:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

row += 1
ws.cell(row, 1, "Revised bottom line, now covering 534-1,000").font = RED_BOLD
row += 1
bottom_line = [
    "No seat-total between 534 and 1,000 makes Indian elections less biased toward whoever",
    "currently leads. The mean winner's bonus is flat across the ENTIRE range. The",
    "incumbent's odds of keeping its majority climb steadily and smoothly all the way to",
    "99.5% at 1,000 seats -- there is no peak, no sweet spot, and no sign the relationship",
    "would reverse even further out. The most 'unbiased' (least entrenching) point in any",
    "range tested so far is simply the SMALLEST seat-total tested -- 534 -- not because it",
    "is fair in absolute terms (it carries the same flat ~1.15pp bonus as every other size),",
    "but because it entrenches the existing lead the LEAST of the options examined.",
]
for line in bottom_line:
    ws.cell(row, 1, line).font = BOLD
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

col_widths = [16, 16, 16, 16, 16, 16, 16, 16]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("seat_bias_analysis.xlsx")
print("Direct Answer sheet rebuilt with 534-1000 data.")
