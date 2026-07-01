#!/usr/bin/env python3
"""
Build Women Candidacy Dashboard Excel Workbook
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

# Paths
REPO_ROOT = Path(__file__).parent.parent.parent
DATA_PROCESSED = REPO_ROOT / "data" / "processed"
OUTPUT_DIR = REPO_ROOT / "coalition-game-theory" / "output"

# Load processed data
state_summary = pd.read_csv(DATA_PROCESSED / "women_candidacy_by_state.csv")
const_summary = pd.read_csv(DATA_PROCESSED / "women_candidacy_by_constituency.csv")
summary_stats = pd.read_csv(DATA_PROCESSED / "women_candidacy_summary.csv").iloc[0].to_dict()

# Sort for better viz
state_summary = state_summary.sort_values('Women_Win_Pct', ascending=False)
const_top = const_summary.nlargest(50, 'Women_Candidacy_Pct')
const_won = const_summary[const_summary['Women_Won'] == True].sort_values('Women_Candidacy_Pct', ascending=False)

# Create workbook
wb = Workbook()
ws_dashboard = wb.active
ws_dashboard.title = "Dashboard"

# ===== DASHBOARD SHEET =====
row = 1
ws_dashboard['A1'] = "Women's Candidacy & Electoral Performance"
ws_dashboard['A1'].font = Font(size=16, bold=True)

row = 3
ws_dashboard[f'A{row}'] = "Summary Statistics"
ws_dashboard[f'A{row}'].font = Font(size=12, bold=True)

row = 4
for i, (key, val) in enumerate(summary_stats.items()):
    ws_dashboard[f'A{row+i}'] = key.replace('_', ' ')
    if 'Pct' in key:
        ws_dashboard[f'B{row+i}'] = f"{val:.2f}%"
    else:
        ws_dashboard[f'B{row+i}'] = f"{int(val):,}"

# Key insights
row = 12
ws_dashboard[f'A{row}'] = "Key Findings"
ws_dashboard[f'A{row}'].font = Font(size=12, bold=True)

insights = [
    f"Women comprise only 0.1% of assembly election candidates ({int(summary_stats['Total_Women_Candidates']):,} / {int(summary_stats['Total_Assembly_Elections_Candidates']):,})",
    f"When women do run, they win ~10% of seats (comparable to overall win rate)",
    f"Only 0.6% of constituencies ({int(summary_stats['Constituencies_With_Women_Winners']):,} / {int(summary_stats['Total_Constituencies']):,}) have elected a woman",
    "Andhra Pradesh leads in women representation and candidacy; several states have zero women candidates",
    "Reserved seats can unlock latent candidate pool: women exist, but aren't fielded",
]

row = 13
for insight in insights:
    ws_dashboard[f'A{row}'] = insight
    ws_dashboard[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws_dashboard.row_dimensions[row].height = 30
    row += 1

# ===== STATE SUMMARY SHEET =====
ws_state = wb.create_sheet("State Summary")
for r_idx, row in enumerate(dataframe_to_rows(state_summary, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_state.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        if isinstance(value, float) and 'Pct' in state_summary.columns[c_idx-1]:
            cell.number_format = '0.00'

# Adjust column widths
ws_state.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws_state.column_dimensions[col].width = 16

# ===== CONSTITUENCIES WITH WOMEN WINNERS =====
ws_won = wb.create_sheet("Women Winners")
for r_idx, row in enumerate(dataframe_to_rows(const_won, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_won.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

ws_won.column_dimensions['A'].width = 20
ws_won.column_dimensions['B'].width = 30
for col in ['C', 'D', 'E']:
    ws_won.column_dimensions[col].width = 15

# ===== TOP CONSTITUENCIES BY WOMEN CANDIDACY =====
ws_top = wb.create_sheet("High Women Candidacy")
for r_idx, row in enumerate(dataframe_to_rows(const_top, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_top.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

ws_top.column_dimensions['A'].width = 20
ws_top.column_dimensions['B'].width = 30
for col in ['C', 'D', 'E']:
    ws_top.column_dimensions[col].width = 15

# ===== README =====
ws_readme = wb.create_sheet("README", 1)
readme_text = """Women's Candidacy Dashboard (Phase 1)

OVERVIEW
This workbook analyzes women's representation in Indian state assembly elections
(1961–2023) to identify candidates for reserved-seat placement under the 33%
Women's Reservation Amendment (2023).

SHEETS

1. Dashboard: Summary statistics and key findings
   - Total women candidates: 481 (0.1% of all)
   - Women's win rate: 10% (48 winners out of 481)
   - Constituencies with women winners: 48 / 8,530 (0.6%)

2. State Summary: Ranked by women's win percentage
   - Women_Candidacy_Pct: % of candidates in state who are women
   - Women_Win_Pct: % of women candidates who won
   - Win_Rate_Gap: women's win rate minus overall win rate (negative = underperformance)
   - Red flags: states with 0% women win rate, high negative gaps

3. Women Winners: Constituencies where women have been elected (48 total)
   - Shows state, constituency name, candidacy rate, confirmation of win

4. High Women Candidacy: Top 50 constituencies by women candidacy rate
   - Target constituencies for reserved-seat placement
   - High candidacy rate suggests latent female candidate supply

INTERPRETATION FOR DELIMITATION

Women comprise ~0.1% of all assembly candidates but win ~10% of contested seats—
suggesting supply constraint, not demand/electability problem.

Andhra Pradesh (2 women winners, high candidacy) and Odisha constituencies show
highest women representation. Reserved seats here would leverage existing candidate pools.

States with 0% women candidates (Manipur, Nagaland, Mizoram, Mysore, Maharashtra)
represent largest opportunity: reservation could unlock new candidate pools.

NEXT STEPS (Phase 2)
- Build electability model: logistic regression on win probability vs. gender/incumbency/state
- Design optimized reserved-seat placement to maximize representation + coalition leverage
- Monte-Carlo state-wise rotation schemes

DATA NOTES
- TCPD Assembly Elections data, 1961–2023
- Only state assembly elections included (Lok Sabha treated separately)
- 483,565 total candidates, 481 women (0.099%)
- 8,530 unique constituencies analyzed
"""

row = 1
for line in readme_text.split('\n'):
    ws_readme[f'A{row}'] = line
    ws_readme[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws_readme.row_dimensions[row].height = 20 if line.strip() and line[0] != ' ' else 15
    row += 1

ws_readme.column_dimensions['A'].width = 100

# Save
output_file = OUTPUT_DIR / "Women_Candidacy_Analysis.xlsx"
wb.save(output_file)
print(f"✅ Dashboard saved: {output_file}")
print(f"\nSheets created:")
print(f"  - Dashboard: Summary statistics + key findings")
print(f"  - State Summary: Ranked by women's win percentage ({len(state_summary)} states)")
print(f"  - Women Winners: Constituencies with elected women ({len(const_won)} constituencies)")
print(f"  - High Women Candidacy: Top 50 constituencies for reserved-seat targeting")
print(f"  - README: Interpretation guide")
