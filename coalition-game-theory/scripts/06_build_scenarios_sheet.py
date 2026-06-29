import pandas as pd
import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("scenario_results.pkl", "rb") as f:
    results = pickle.load(f)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
BLUE_INPUT = Font(name=FONT_NAME, size=10, color="0000FF")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HUNG_FILL = PatternFill("solid", start_color="F8CBAD")
MAJORITY_FILL = PatternFill("solid", start_color="C6E0B4")

wb = load_workbook("coalition_model.xlsx")
ws = wb.create_sheet("Scenarios")
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Scenario Definitions: Hypothetical Seat Distributions (existing parties only)").font = TITLE_FONT
ws.merge_cells("A1:N1")
ws.cell(2, 1, "Edit blue cells to test your own assumptions. Majority threshold = 272 of 543 seats.").font = NOTE_FONT
ws.merge_cells("A2:N2")

scenario_order = ["S0_Actual_Current","S1_2024_AsDeclared","S2_NDA_Falls_Short",
                  "S3_Deep_Hung_Parliament","S4_NDA_Needs_Swing","S5_INDIA_Largest_No_Majority"]
scenario_labels = {
    "S0_Actual_Current": "S0: Actual Current (June 2026)",
    "S1_2024_AsDeclared": "S1: 2024 Result As Declared",
    "S2_NDA_Falls_Short": "S2: NDA Falls Short, Hung",
    "S3_Deep_Hung_Parliament": "S3: Deep Fragmentation",
    "S4_NDA_Needs_Swing": "S4: NDA Needs One Swing Party",
    "S5_INDIA_Largest_No_Majority": "S5: INDIA Largest, No Majority",
}

row = 4
all_party_names = []
for s in scenario_order:
    for p, w in results[s]["players"]:
        if p not in all_party_names:
            all_party_names.append(p)

# Build a table: rows = parties, columns = scenarios
headers = ["Party / Bloc"] + [scenario_labels[s] for s in scenario_order]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

start_row = row + 1
for i, party in enumerate(all_party_names):
    r = start_row + i
    ws.cell(r, 1, party).font = NORMAL
    ws.cell(r, 1).border = BORDER
    for j, s in enumerate(scenario_order):
        pw = dict(results[s]["players"])
        val = pw.get(party, 0)
        c = ws.cell(r, j+2, val if val else None)
        c.font = BLUE_INPUT
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

last_row = start_row + len(all_party_names) - 1

# Totals row
trow = last_row + 1
ws.cell(trow, 1, "TOTAL").font = BOLD
ws.cell(trow, 1).border = BORDER
for j, s in enumerate(scenario_order):
    col = get_column_letter(j+2)
    ws.cell(trow, j+2, f"=SUM({col}{start_row}:{col}{last_row})").font = BOLD
    ws.cell(trow, j+2).border = BORDER
    ws.cell(trow, j+2).fill = PatternFill("solid", start_color="D9E1F2")

# Majority status row
mrow = trow + 1
ws.cell(mrow, 1, "RESULT").font = BOLD
for j, s in enumerate(scenario_order):
    is_hung = results[s]["is_hung"]
    label = "HUNG PARLIAMENT" if is_hung else "Single bloc has majority"
    c = ws.cell(mrow, j+2, label)
    c.font = BOLD
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.fill = HUNG_FILL if is_hung else MAJORITY_FILL
    c.border = BORDER

ws.row_dimensions[mrow].height = 30

col_widths = [22] + [20]*len(scenario_order)
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "B5"

wb.save("coalition_model.xlsx")
print(f"Scenarios sheet done. Parties: {len(all_party_names)}, rows {start_row}-{last_row}")
