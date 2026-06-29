import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
NORMAL = Font(name=FONT_NAME, size=9)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = load_workbook("election_ml_tcpd_real.xlsx")
df = pd.read_csv("ga_features.csv")

ws = wb.create_sheet("Raw PC Winners Sample")
ws.sheet_view.showGridLines = False
ws.cell(1, 1, "Sample of Rolled-Up PC-Level Winner Data (real, post-cleaning)").font = TITLE_FONT
ws.merge_cells("A1:L1")
ws.cell(2, 1, "Showing sample rows across all 5 elections for transparency/spot-checking. Full 2,704-row dataset used for modeling.").font = NOTE_FONT
ws.merge_cells("A2:L2")

cols = ["State_Name","Year","PC_Name","PC_No","Party","Bloc","Candidate","Total_Votes",
        "Margin_Pct","N_Candidates_PC","Effective_N_Parties","split"]
headers = ["State","Year","PC Name","PC No","Party","Bloc","Candidate (Winner)","Votes","Margin %","N Candidates","Eff. N Parties","Split"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(4, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

sample = df.sort_values(["Year","State_Name","PC_No"]).groupby("Year").head(40).reset_index(drop=True)
start_row = 5
for i, row in sample.iterrows():
    r = start_row + i
    for c, col in enumerate(cols, start=1):
        val = row[col]
        if col in ("Margin_Pct", "Effective_N_Parties") and pd.notna(val):
            val = round(float(val), 2)
        ws.cell(r, c, val if pd.notna(val) else None).font = NORMAL
        ws.cell(r, c).border = BORDER
last_row = start_row + len(sample) - 1

col_widths = [16, 8, 16, 8, 10, 10, 22, 10, 10, 12, 12, 10]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

wb.save("election_ml_tcpd_real.xlsx")
print(f"Raw sample sheet done. rows {start_row}-{last_row}")
