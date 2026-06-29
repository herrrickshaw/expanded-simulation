import pandas as pd
import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

with open("scenario_results.pkl", "rb") as f:
    results = pickle.load(f)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = load_workbook("coalition_model.xlsx")
ws = wb.create_sheet("Banzhaf Power Index")
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Banzhaf Power Index by Scenario").font = TITLE_FONT
ws.merge_cells("A1:H1")
ws.cell(2, 1, "Share of all possible coalitions in which each party is the decisive (pivotal) vote. A party can have real bargaining power even with few seats — or none, even with many — depending on the configuration.").font = NOTE_FONT
ws.merge_cells("A2:H2")

scenario_order = ["S0_Actual_Current","S1_2024_AsDeclared","S2_NDA_Falls_Short",
                  "S3_Deep_Hung_Parliament","S4_NDA_Needs_Swing","S5_INDIA_Largest_No_Majority"]
scenario_labels = {
    "S0_Actual_Current": "S0: Current",
    "S1_2024_AsDeclared": "S1: 2024 Declared",
    "S2_NDA_Falls_Short": "S2: NDA Short",
    "S3_Deep_Hung_Parliament": "S3: Fragmented",
    "S4_NDA_Needs_Swing": "S4: NDA+Swing",
    "S5_INDIA_Largest_No_Majority": "S5: INDIA Largest",
}

all_parties = []
for s in scenario_order:
    for p, b in results[s]["banzhaf"]:
        if p not in all_parties:
            all_parties.append(p)

row = 4
headers = ["Party / Bloc"] + [scenario_labels[s] for s in scenario_order]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

start_row = row + 1
# Order parties: NDA, INDIA first, then by max banzhaf descending across scenarios
def max_banzhaf(party):
    vals = []
    for s in scenario_order:
        d = dict(results[s]["banzhaf"])
        vals.append(d.get(party, 0))
    return max(vals)

ordered_parties = sorted(all_parties, key=lambda p: (-1 if p in ("NDA","INDIA") else 0, -max_banzhaf(p)))

for i, party in enumerate(ordered_parties):
    r = start_row + i
    ws.cell(r, 1, party).font = NORMAL
    ws.cell(r, 1).border = BORDER
    for j, s in enumerate(scenario_order):
        d = dict(results[s]["banzhaf"])
        val = d.get(party, None)
        c = ws.cell(r, j+2, round(val, 3) if val is not None else None)
        c.font = NORMAL
        c.number_format = "0.0%"
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

last_row = start_row + len(ordered_parties) - 1

col_widths = [16] + [16]*len(scenario_order)
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "B5"

# Chart: Banzhaf index across scenarios for top parties (NDA, INDIA, + top 2 swing parties)
chart = BarChart()
chart.type = "col"
chart.title = "Banzhaf Power Index: NDA vs INDIA vs Key Swing Parties, by Scenario"
chart.y_axis.title = "Banzhaf Power Index"
chart.x_axis.title = "Scenario"

# Build small data block for chart: rows = scenario, cols = NDA, INDIA, YSRCP (most consistently relevant swing party)
chart_row = last_row + 3
ws.cell(chart_row, 1, "Scenario").font = BOLD
ws.cell(chart_row, 2, "NDA").font = BOLD
ws.cell(chart_row, 3, "INDIA").font = BOLD
ws.cell(chart_row, 4, "YSRCP").font = BOLD
for c in range(1, 5):
    ws.cell(chart_row, c).fill = HEADER_FILL
    ws.cell(chart_row, c).font = HEADER_FONT
    ws.cell(chart_row, c).border = BORDER

for j, s in enumerate(scenario_order):
    r = chart_row + 1 + j
    d = dict(results[s]["banzhaf"])
    ws.cell(r, 1, scenario_labels[s]).font = NORMAL
    ws.cell(r, 2, round(d.get("NDA", 0), 3)).font = NORMAL
    ws.cell(r, 3, round(d.get("INDIA", 0), 3)).font = NORMAL
    ws.cell(r, 4, round(d.get("YSRCP", 0), 3)).font = NORMAL
    for c in range(1, 5):
        ws.cell(r, c).border = BORDER
        if c > 1:
            ws.cell(r, c).number_format = "0.0%"

chart_last_row = chart_row + len(scenario_order)

data_ref = Reference(ws, min_col=2, max_col=4, min_row=chart_row, max_row=chart_last_row)
cats_ref = Reference(ws, min_col=1, min_row=chart_row+1, max_row=chart_last_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 10
chart.width = 20
ws.add_chart(chart, f"A{chart_last_row+3}")

wb.save("coalition_model.xlsx")
print("Banzhaf sheet done.")
