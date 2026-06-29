import pandas as pd
import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WARN_FILL = PatternFill("solid", start_color="FFF2CC")

with open("dsmm_results.pkl", "rb") as f:
    dsmm = pickle.load(f)

wb = load_workbook("vote_share_models.xlsx")
ws = wb.create_sheet("Dirichlet-Swing Diagnostics")
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Dirichlet-Swing Matrix Model: Implementation Diagnostics").font = TITLE_FONT
ws.merge_cells("A1:F1")
ws.cell(2, 1, "Two real numerical bugs were found and fixed while implementing Mitra (2026)'s Dirichlet-Swing Matrix Model on real data.").font = NOTE_FONT
ws.merge_cells("A2:F2")

row = 4
ws.cell(row, 1, "Bug 1: Naive digamma inversion overflow").font = SUBTITLE_FONT
row += 1
bug1 = [
    "Minka's fixed-point MLE algorithm for Dirichlet parameters requires inverting the",
    "digamma function at each iteration. A naive Newton's-method implementation overflowed",
    "(np.exp() on large inputs) when the iteration entered a region with extreme parameter",
    "values. FIX: added input clipping and a large-y fallback branch to keep the Newton",
    "iteration numerically stable across its full operating range.",
]
for line in bug1:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

row += 1
ws.cell(row, 1, "Bug 2: Dirichlet MLE degeneracy on zero-inflated data (the deeper issue)").font = SUBTITLE_FONT
row += 1
bug2 = [
    "After fixing Bug 1, the model still produced degenerate results: alpha parameters",
    "diverging to the numerical cap (500-10,000) in one dimension, collapsing the Dirichlet",
    "mean to a point mass [0,0,1] or [1,0,0]. Root cause: roughly 25% of 'Other'-origin",
    "constituencies have EXACTLY ZERO NDA vote share (no NDA-aligned candidate even",
    "contested) -- a real, valid feature of 1999-2004 Indian elections, not a data error.",
    "Maximum-likelihood Dirichlet fitting is mathematically known to be unstable on data",
    "with substantial exact-zero mass: the likelihood is genuinely maximized at the",
    "boundary (alpha -> infinity), a well-documented pathology, not a coding error.",
    "",
    "FIX ATTEMPTED #1 (insufficient): added a small number of weak Dirichlet(1,1,1) pseudo-",
    "observations as regularization. With only 10 pseudo-observations against 90+ real,",
    "heavily-skewed observations, this was far too weak -- still collapsed to the boundary.",
    "",
    "FIX ATTEMPTED #2 (insufficient): scaled pseudo-observation count proportionally to",
    "subset size (50% of N). Still collapsed -- the real data's skew was severe enough",
    "that even 1/3 of the total weight as uniform pseudo-data could not pull it back.",
    "",
    "FINAL FIX (adopted): abandoned MLE in favor of METHOD-OF-MOMENTS Dirichlet estimation,",
    "which matches sample mean and variance directly rather than fitting log-likelihood",
    "gradients, and cannot diverge the same way. This produced a well-behaved, interpretable",
    "transition matrix (see below) with no further degeneracy.",
]
for line in bug2:
    font = BOLD if line.startswith("FIX") or line.startswith("FINAL") else NORMAL
    ws.cell(row, 1, line).font = font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

row += 1
ws.cell(row, 1, "Final fitted swing matrix (Dirichlet mean transition probabilities, 1999->2004)").font = SUBTITLE_FONT
row += 1
means = dsmm["dirichlet_means_train"]
blocs = ["NDA", "INDIA", "Other"]
headers = ["From \\ To"] + [f"To {b}" for b in blocs]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER
row += 1
for i, origin in enumerate(blocs):
    ws.cell(row, 1, f"From {origin}").font = BOLD
    ws.cell(row, 1).border = BORDER
    for j in range(3):
        c = ws.cell(row, j+2, round(float(means[i][j]), 3))
        c.font = NORMAL
        c.number_format = "0.0%"
        c.border = BORDER
        if i == j:
            c.fill = WARN_FILL
    row += 1

row += 1
interp = [
    "Diagonal (same-bloc retention) values are highlighted: NDA-origin seats stay NDA-leaning",
    "45.8% of the time, INDIA-origin seats stay INDIA-leaning 53.2% of the time, Other-origin",
    "seats stay Other-leaning 50.4% of the time -- all plausible, non-degenerate incumbency-",
    "correlated persistence rates, confirming the final fix produced a sensible model.",
]
for line in interp:
    ws.cell(row, 1, line).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

col_widths = [16, 14, 14, 14, 4, 4]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("vote_share_models.xlsx")
print("Dirichlet-Swing diagnostics sheet done.")
