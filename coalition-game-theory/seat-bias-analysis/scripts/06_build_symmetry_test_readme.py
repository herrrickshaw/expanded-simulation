import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
RED_BOLD = Font(name=FONT_NAME, bold=True, size=10, color="C00000")

wb = Workbook()
ws0 = wb.active
ws0.title = "README"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 112

r = 1
ws0.cell(r, 1, "Counterfactual Scenario Test: Does Entrenchment Favor NDA Specifically, or Whoever Leads?").font = TITLE_FONT
r += 2

notes = [
    ("On the word 'GAN'", BOLD),
    ("  A literal Generative Adversarial Network (a generator/discriminator pair trained via", NORMAL),
    ("  adversarial loss) needs real training data to learn a distribution from. Our 'data' is", NORMAL),
    ("  itself a parametric simulation (Dirichlet draws around state-level means), not an", NORMAL),
    ("  empirical sample large enough to train a GAN against. What's built here is a", NORMAL),
    ("  COUNTERFACTUAL SCENARIO GENERATOR: the same validated simulation engine, re-run under", NORMAL),
    ("  different lead-favoring configurations, so results are directly comparable.", NORMAL),
    ("", NORMAL),
    ("Five scenarios tested", BOLD),
    ("  S0 NDA Real      -- NDA leads by its actual 2024 margin (+11.1pp)", NORMAL),
    ("  S1 INDIA Mirror  -- exact mirror: INDIA leads by the SAME +11.1pp margin", NORMAL),
    ("  S2 Tossup        -- genuine 50/50 race between NDA and INDIA (the literal fair coin)", NORMAL),
    ("  S3 INDIA Small   -- INDIA leads by a small +3pp margin", NORMAL),
    ("  S4 NDA Small     -- NDA leads by a small +3pp margin (mirror of S3)", NORMAL),
    ("  All five use the same simulation engine validated in the prior 534-1,000 seat sweep,", NORMAL),
    ("  varying ONLY which bloc leads and by how much.", NORMAL),
    ("", NORMAL),
    ("Two real issues found and resolved while building this -- documented for transparency", RED_BOLD),
    ("  ISSUE 1 (a genuine bug, fixed): the first scenario-construction method shifted", NORMAL),
    ("  NDA/INDIA shares ADDITIVELY (nda+shift, india-shift), which could push values below", NORMAL),
    ("  zero before clipping when flipping an 11pp lead to the opposite side (a ~22pp total", NORMAL),
    ("  swing). The clip distorted S0 and S1 asymmetrically -- they were NOT exact mirrors", NORMAL),
    ("  despite being intended as such (53.1% vs 51.8% tracked vote share).", NORMAL),
    ("  A logit-space attempt got closer but was still off by ~2pp, because each state's", NORMAL),
    ("  'Other' share differs, so a uniform logit shift didn't produce a symmetric NATIONAL", NORMAL),
    ("  aggregate. FINAL FIX: abandoned per-state regional-lean preservation in favor of a", NORMAL),
    ("  UNIFORM national NDA-vs-INDIA split (each state's real 'Other' share kept, but the", NORMAL),
    ("  NDA/INDIA contest made uniform nationally). This guarantees exact symmetry by", NORMAL),
    ("  construction -- verified to 6 decimal places before trusting any further results.", NORMAL),
    ("", NORMAL),
    ("  ISSUE 2 (looked like a bug, was actually correct): the toss-up scenario (S2) showed", NORMAL),
    ("  NDA's seat share (48.7%) diverging from its vote share (47.6%), and P(NDA majority)", NORMAL),
    ("  FALLING as seats increased. Investigation showed this is correct: with NDA and INDIA", NORMAL),
    ("  tied at ~48.7% each (the rest going to 'Other'), NEITHER bloc clears 50% on average,", NORMAL),
    ("  so more seats makes a HUNG outcome more likely for BOTH sides symmetrically (verified:", NORMAL),
    ("  P(NDA win)=26.0%, P(INDIA win)=25.9%, nearly identical). This is the SAME entrenchment", NORMAL),
    ("  mechanism applied to a hung outcome instead of a majority -- not a bug.", NORMAL),
    ("", NORMAL),
    ("Sources", BOLD),
    ("  Builds on the validated 534-1,000 seat sweep methodology from this project's earlier", NORMAL),
    ("  'Seat-Total Bias Analysis' workbook. Real 2024 state-level NDA/INDIA/Other bloc data.", NORMAL),
]
for text, font in notes:
    ws0.cell(r, 1, text).font = font
    r += 1

wb.save("gan_symmetry_test.xlsx")
print("README done.")
