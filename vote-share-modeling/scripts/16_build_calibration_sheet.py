import pandas as pd
import pickle
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
RED_BOLD = Font(name=FONT_NAME, bold=True, size=10, color="C00000")
NORMAL = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WARN_FILL = PatternFill("solid", start_color="F8CBAD")

with open("hierarchical_model_results.pkl", "rb") as f:
    hier = pickle.load(f)

wb = load_workbook("vote_share_models.xlsx")
ws = wb.create_sheet("Hierarchical Model Calibration")
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Hierarchical Bayesian Model: Credible Interval Calibration Check").font = TITLE_FONT
ws.merge_cells("A1:H1")
ws.cell(2, 1, "Tests whether the model's stated uncertainty (90% credible intervals) is honest -- whether ~90% of actual test outcomes fall inside the predicted interval.").font = NOTE_FONT
ws.merge_cells("A2:H2")

row = 4
ws.cell(row, 1, f"RESULT: 90% credible intervals achieved only {hier['coverage']:.1%} actual coverage on test").font = RED_BOLD
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1
ws.cell(row, 1, "(A well-calibrated model would show ~90%. 13.5% means the model is badly OVERCONFIDENT on test data.)").font = NOTE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 2

ws.cell(row, 1, "Why this happened").font = SUBTITLE_FONT
row += 1
why = [
    "Mean NDA vote share genuinely shifted from 0.331 (1999-2004, train era) to 0.407",
    "(2014-2019, test era) -- a real national +7.6 percentage point swing. The model's",
    "alpha_national parameter was fit ONLY on train-era data, so its posterior is",
    "appropriately narrow for IN-REGIME prediction, but has zero information about a",
    "national-level shift occurring after the training window ends. This is not a bug in",
    "the model's math -- it is an honest, correct reflection of what the training data",
    "could possibly tell it. No amount of better MCMC sampling would fix this; only",
    "information about the FUTURE shift itself (e.g. contemporaneous polling) could.",
]
for line in why:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

row += 1
ws.cell(row, 1, "Sample of test-set predictions vs actuals (first 25 rows)").font = SUBTITLE_FONT
row += 1
headers = ["State","Year","PC Name","Predicted V_NDA","90% CI Lower","90% CI Upper","Actual V_NDA","Inside CI?"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER
row += 1
start_row = row
n_show = 25
for i in range(min(n_show, len(hier["test_pred"]))):
    ws.cell(row, 1, hier["test_states"][i]).font = NORMAL
    ws.cell(row, 2, int(hier["test_years"][i])).font = NORMAL
    ws.cell(row, 3, hier["test_pc_names"][i]).font = NORMAL
    ws.cell(row, 4, round(float(hier["test_pred"][i]), 3)).font = NORMAL
    ws.cell(row, 5, round(float(hier["test_lo"][i]), 3)).font = NORMAL
    ws.cell(row, 6, round(float(hier["test_hi"][i]), 3)).font = NORMAL
    ws.cell(row, 7, round(float(hier["test_actual"][i]), 3)).font = BOLD
    inside = hier["test_lo"][i] <= hier["test_actual"][i] <= hier["test_hi"][i]
    c8 = ws.cell(row, 8, "Yes" if inside else "No")
    c8.font = BOLD
    if not inside:
        c8.fill = WARN_FILL
    for c in range(1, 9):
        ws.cell(row, c).border = BORDER
    row += 1
last_row = row - 1

col_widths = [20, 8, 22, 16, 12, 12, 14, 10]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = f"A{start_row}"

wb.save("vote_share_models.xlsx")
print(f"Calibration sheet done. rows {start_row}-{last_row}")
