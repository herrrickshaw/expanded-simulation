import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, color="404040")
BOLD = Font(name=FONT_NAME, bold=True, size=11)
NORMAL = Font(name=FONT_NAME, size=10)
BIG_NUM = Font(name=FONT_NAME, bold=True, size=20, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CARD_FILL = PatternFill("solid", start_color="EAF1F8")

wb = load_workbook("election_analysis.xlsx")
df = pd.read_csv("state_data_raw.csv")
df["AvgPerSeat"] = df["Electors_2024"] / df["LS_Seats_2024"]
df_states = df.dropna(subset=["AvgPerSeat"]).sort_values("AvgPerSeat", ascending=False).reset_index(drop=True)

ws = wb.create_sheet("Dashboard", 0)  # make it first sheet
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "India Elections: Average Voters per Constituency — Dashboard").font = TITLE_FONT
ws.merge_cells("A1:H1")
ws.cell(2, 1, "Lok Sabha (2024) state-wise comparison, with population-growth context | Compiled June 2026").font = SUBTITLE_FONT
ws.merge_cells("A2:H2")

# KPI cards
kpis = [
    ("All-India Avg Electors/LS Seat (2024)", "='State-wise LS Data'!D39", "#,##0"),
    ("Highest State Avg Electors/Seat (Delhi)", "='State-wise LS Data'!D21", "#,##0"),
    ("Total LS Seats (frozen since 1976)", "='State-wise LS Data'!B39", "0"),
    ("Growth in Avg Electors/Seat, 2014→2024", "='National Trend'!B9", "0.0%"),
]
col = 1
for label, formula, fmt in kpis:
    ws.cell(4, col, label).font = NORMAL
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    c = ws.cell(5, col, formula)
    c.font = BIG_NUM
    c.number_format = fmt
    ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col+1)
    for rr in range(4, 7):
        for cc in range(col, col+2):
            ws.cell(rr, cc).fill = CARD_FILL
            ws.cell(rr, cc).border = BORDER
    col += 2

# Top 10 / Bottom 10 table for chart
ws.cell(8, 1, "Top 10 states by average electors per Lok Sabha constituency (2024)").font = BOLD
top10 = df_states.head(10)[["State", "AvgPerSeat"]]
hdr_r = 9
ws.cell(hdr_r, 1, "State").font = HEADER_FONT
ws.cell(hdr_r, 1).fill = HEADER_FILL
ws.cell(hdr_r, 2, "Avg Electors/Seat").font = HEADER_FONT
ws.cell(hdr_r, 2).fill = HEADER_FILL
for i, row in top10.iterrows():
    r = hdr_r + 1 + i
    ws.cell(r, 1, row["State"]).font = NORMAL
    ws.cell(r, 2, round(row["AvgPerSeat"])).font = NORMAL
    ws.cell(r, 2).number_format = "#,##0"
    for c in (1, 2):
        ws.cell(r, c).border = BORDER
top10_last = hdr_r + len(top10)

ws.cell(8, 4, "Bottom 10 states by average electors per Lok Sabha constituency (2024)").font = BOLD
bottom10 = df_states.tail(10).sort_values("AvgPerSeat")[["State", "AvgPerSeat"]]
ws.cell(hdr_r, 4, "State").font = HEADER_FONT
ws.cell(hdr_r, 4).fill = HEADER_FILL
ws.cell(hdr_r, 5, "Avg Electors/Seat").font = HEADER_FONT
ws.cell(hdr_r, 5).fill = HEADER_FILL
for i, row in bottom10.reset_index(drop=True).iterrows():
    r = hdr_r + 1 + i
    ws.cell(r, 4, row["State"]).font = NORMAL
    ws.cell(r, 5, round(row["AvgPerSeat"])).font = NORMAL
    ws.cell(r, 5).number_format = "#,##0"
    for c in (4, 5):
        ws.cell(r, c).border = BORDER
bottom10_last = hdr_r + len(bottom10)

# Charts
chart1 = BarChart()
chart1.type = "bar"
chart1.title = "Top 10: Highest avg electors per LS seat"
data_ref = Reference(ws, min_col=2, min_row=hdr_r, max_row=top10_last)
cats_ref = Reference(ws, min_col=1, min_row=hdr_r+1, max_row=top10_last)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.height = 9
chart1.width = 16
ws.add_chart(chart1, "A21")

chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Bottom 10: Lowest avg electors per LS seat"
data_ref2 = Reference(ws, min_col=5, min_row=hdr_r, max_row=bottom10_last)
cats_ref2 = Reference(ws, min_col=4, min_row=hdr_r+1, max_row=bottom10_last)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
chart2.height = 9
chart2.width = 16
ws.add_chart(chart2, "I21")

# Key takeaways box
tr = 40
ws.cell(tr, 1, "Key takeaways").font = BOLD
takeaways = [
    "1. Constituency size varies more than 6x across states — small/hill states (Lakshadweep, Sikkim, NE states) have under",
    "   1 lakh to ~5 lakh electors per seat, while large states' urban seats (Telangana's Malkajgiri, Karnataka's Bangalore North,",
    "   Maharashtra's Mumbai/Pune belt) exceed 30 lakh — a direct legacy of the 1976/2002 seat freeze not tracking growth.",
    "2. National average electors per LS seat rose from ~15.4 lakh (2014) to ~17.8 lakh (2024) — about 16% growth in a decade —",
    "   while the total seat count stayed at 543. Population is growing; representation slots are not.",
    "3. States with strong family-planning-driven population deceleration (e.g. southern states) now have smaller, denser-served",
    "   constituencies than fast-growing northern states (UP, Bihar) — a long-standing structural imbalance the freeze created.",
    "4. The Delimitation Bill, 2026 (introduced in Parliament 16 April 2026) proposes resetting the population basis to the 2011",
    "   Census and raising the Lok Sabha's maximum size from 550 to 850 seats — directly responding to the dilution this dashboard",
    "   quantifies. If enacted, expect a major reallocation toward higher-population states.",
]
for i, line in enumerate(takeaways):
    ws.cell(tr+1+i, 1, line).font = NORMAL
    ws.merge_cells(start_row=tr+1+i, start_column=1, end_row=tr+1+i, end_column=8)

col_widths = [24, 16, 4, 24, 16, 4, 4, 4]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("election_analysis.xlsx")
print("Dashboard sheet done.")
print(wb.sheetnames)
