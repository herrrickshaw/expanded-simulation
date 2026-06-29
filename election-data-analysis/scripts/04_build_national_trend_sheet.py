import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
BLUE_INPUT = Font(name=FONT_NAME, size=10, color="0000FF")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = load_workbook("election_analysis.xlsx")
ndf = pd.read_csv("national_trend.csv")
df = pd.read_csv("state_data_raw.csv")
df["is_state"] = df["Assembly_Seats"].notna()
df = df.sort_values(["is_state","LS_Seats_2024"], ascending=[False, False]).reset_index(drop=True)

# ============================================================
# SHEET 3: National Trend
# ============================================================
ws2 = wb.create_sheet("National Trend")
ws2.sheet_view.showGridLines = False
ws2.cell(1, 1, "National Lok Sabha Trend: Electors, Turnout & Seats (2014–2024)").font = TITLE_FONT
ws2.merge_cells("A1:F1")

headers = ["Election Year", "Total Registered Electors", "Turnout (%)", "Votes Cast", "LS Seats Contested", "Avg Electors per Seat"]
for c, h in enumerate(headers, start=1):
    cell = ws2.cell(3, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

start_row = 4
for i, row in ndf.iterrows():
    r = start_row + i
    ws2.cell(r, 1, int(row["Year"])).font = BLUE_INPUT
    ws2.cell(r, 2, int(row["Total_Electors"])).font = BLUE_INPUT
    ws2.cell(r, 2).number_format = "#,##0"
    ws2.cell(r, 3, float(row["Turnout_Pct"])).font = BLUE_INPUT
    ws2.cell(r, 3).number_format = "0.00"
    ws2.cell(r, 4, int(row["Votes_Cast"])).font = BLUE_INPUT
    ws2.cell(r, 4).number_format = "#,##0"
    ws2.cell(r, 5, int(row["LS_Seats"])).font = BLUE_INPUT
    ws2.cell(r, 6, f"=B{r}/E{r}").font = NORMAL
    ws2.cell(r, 6).number_format = "#,##0"
    for c in range(1, 7):
        ws2.cell(r, c).border = BORDER

last_row = start_row + len(ndf) - 1

gr = last_row + 2
ws2.cell(gr, 1, "Growth 2014→2024 (electors):").font = BOLD
ws2.cell(gr, 2, f"=(B{last_row}-B{start_row})/B{start_row}").font = NORMAL
ws2.cell(gr, 2).number_format = "0.0%"
ws2.cell(gr+1, 1, "Growth 2014→2024 (avg electors/seat):").font = BOLD
ws2.cell(gr+1, 2, f"=(F{last_row}-F{start_row})/F{start_row}").font = NORMAL
ws2.cell(gr+1, 2).number_format = "0.0%"
ws2.cell(gr+2, 1, "LS seats, 2014 vs 2024:").font = BOLD
ws2.cell(gr+2, 2, f"=E{last_row}-E{start_row}").font = NORMAL
ws2.cell(gr+3, 1, "Interpretation:").font = BOLD
ws2.cell(gr+3, 2, "Electorate grew ~16% per seat in a decade while seat count stayed essentially flat — representation is diluting.").font = NORMAL
ws2.merge_cells(start_row=gr+3, start_column=2, end_row=gr+3, end_column=6)

note_r = gr + 5
ws2.cell(note_r, 1, "Note: 2014/2019 figures are rounded approximations from ECI/PIB releases pending full constituency-level reconciliation;").font = NOTE_FONT
ws2.cell(note_r+1, 1, "2024 figure (96.88cr) is the EC's officially announced final electoral roll total. See README for sourcing.").font = NOTE_FONT

col_widths = [16, 24, 14, 18, 16, 18]
for i, w in enumerate(col_widths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Chart: Avg electors per seat trend (line)
chart1 = LineChart()
chart1.title = "Average Electors per Lok Sabha Seat (2014–2024)"
chart1.style = 2
chart1.y_axis.title = "Avg Electors per Seat"
chart1.x_axis.title = "Election Year"
data_ref = Reference(ws2, min_col=6, min_row=3, max_row=last_row)
cats_ref = Reference(ws2, min_col=1, min_row=start_row, max_row=last_row)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.height = 8
chart1.width = 16
ws2.add_chart(chart1, "A12")

wb.save("election_analysis.xlsx")
print("Sheet 3 done. rows:", start_row, last_row)
