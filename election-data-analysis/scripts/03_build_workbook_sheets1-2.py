import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.comments import Comment

# ---------- Load data ----------
df = pd.read_csv("state_data_raw.csv")
ndf = pd.read_csv("national_trend.csv")

# Sort states by LS seats descending for readability; keep UTs at bottom
df["is_state"] = df["Assembly_Seats"].notna()
df = df.sort_values(["is_state","LS_Seats_2024"], ascending=[False, False]).reset_index(drop=True)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="595959")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
BLUE_INPUT = Font(name=FONT_NAME, size=10, color="0000FF")
GREEN_LINK = Font(name=FONT_NAME, size=10, color="008000")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YELLOW_FLAG = PatternFill("solid", start_color="FFF2CC")

wb = Workbook()

# ============================================================
# SHEET 1: README / Methodology
# ============================================================
ws0 = wb.active
ws0.title = "README"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 110

r = 1
ws0.cell(r, 1, "India Election Data: Voters per Constituency — State-wise Analysis & Growth Trends").font = TITLE_FONT
r += 2
ws0.cell(r, 1, "Scope: Lok Sabha (2014, 2019, 2024) + State Legislative Assemblies (current seat allocation), all-India, state-wise comparison.").font = NORMAL
r += 2

notes = [
    ("Sheet guide", BOLD),
    ("  1. State-wise LS Data — 2024 electors, seats, 2011 Census population, Assembly seats per state, with calculated averages.", NORMAL),
    ("  2. National Trend — total electors, turnout and average electors/seat for 2014, 2019, 2024 Lok Sabha elections.", NORMAL),
    ("  3. Growth & Delimitation — population-per-seat using 1971 vs 2011 Census, flagging states most under/over-represented.", NORMAL),
    ("  4. Dashboard — charts summarizing the above.", NORMAL),
    ("", NORMAL),
    ("Key methodology notes", BOLD),
    ("  • 'Average voters per constituency' = state's total registered electors (2024) ÷ number of Lok Sabha seats allotted to that state.", NORMAL),
    ("  • Electors_2024 figures are aggregated by summing official constituency-level elector counts (ECI 2024 statistical reports, as", NORMAL),
    ("    compiled in Wikipedia's 'List of constituencies of the Lok Sabha'). The sum of all states (97.36 crore) is within ~0.5% of the", NORMAL),
    ("    Election Commission's officially announced final electoral roll figure of 96.88 crore — small variance likely reflects rounding", NORMAL),
    ("    and the gap between the pre-poll published roll (cited nationally) and final constituency-level rolls used on polling day.", NORMAL),
    ("    Treat state-level figures as accurate to within roughly 1% — suitable for comparative ranking, not for legal/statutory use.", NORMAL),
    ("  • LS seats per state have been FROZEN since 1976 (based on 1971 Census) and remain frozen at the 2001-Census delimitation of", NORMAL),
    ("    boundaries until after the first Census post-2026. This is the single most important structural fact behind this analysis:", NORMAL),
    ("    seat counts do NOT track current population, so 'voters per constituency' varies enormously by state.", NORMAL),
    ("  • A 2026 Delimitation Bill is currently before Parliament (introduced 16 April 2026) proposing to base future delimitation on", NORMAL),
    ("    the 2011 Census (not 1971) and raise the maximum Lok Sabha size from 550 to 850 seats. This would directly address the", NORMAL),
    ("    imbalances quantified in this workbook. Source: PRS Legislative Research, accessed June 2026.", NORMAL),
    ("  • Assembly (Vidhan Sabha) seat counts are current allocations per state, used for cross-check of national vs state-level", NORMAL),
    ("    representation density.", NORMAL),
    ("", NORMAL),
    ("Sources", BOLD),
    ("  • Election Commission of India — 2024 Lok Sabha statistical reports (constituency-wise electors)", NORMAL),
    ("  • Census of India 2011 — state population (basis for current constituency delimitation)", NORMAL),
    ("  • Wikipedia: 'List of constituencies of the Lok Sabha', 'Lok Sabha', 'Delimitation Commission of India' (compiled ECI/Census data)", NORMAL),
    ("  • PRS Legislative Research — The Delimitation Bill, 2026 (introduced 16 April 2026)", NORMAL),
    ("  • PIB / BusinessToday — national elector totals 2014/2019/2024", NORMAL),
    ("", NORMAL),
    ("Caveat", BOLD),
    ("  This is a desk compilation from public sources for analytical/comparative purposes, not an official ECI publication.", NORMAL),
    ("  Figures for 2014/2019 national totals are approximate (rounded to nearest 0.1 crore) pending full constituency-level", NORMAL),
    ("  reconciliation; treat the 2024 state-wise breakdown as the most granular and reliable layer in this workbook.", NORMAL),
]
for text, font in notes:
    ws0.cell(r, 1, text).font = font
    r += 1

# ============================================================
# SHEET 2: State-wise LS Data
# ============================================================
ws1 = wb.create_sheet("State-wise LS Data")
ws1.sheet_view.showGridLines = False

headers = ["State", "LS Seats (2024)", "Electors (2024)", "Avg Electors/Seat",
           "2011 Census Population", "Pop per Seat (2011)", "Assembly Seats",
           "Avg Pop/Assembly Seat", "LS:Assembly Seat Ratio"]
ws1.append(["" for _ in headers])  # placeholder row1 for title merge later
for c, h in enumerate(headers, start=1):
    cell = ws1.cell(2, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

ws1.cell(1, 1, "State-wise Lok Sabha Constituency & Population Data (2024 election / 2011 Census)").font = TITLE_FONT
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

start_row = 3
for i, row in df.iterrows():
    r = start_row + i
    ws1.cell(r, 1, row["State"]).font = NORMAL
    ws1.cell(r, 2, int(row["LS_Seats_2024"])).font = BLUE_INPUT
    elec = row["Electors_2024"]
    ws1.cell(r, 3, int(elec) if pd.notna(elec) else None).font = BLUE_INPUT
    ws1.cell(r, 4, f"=IF(C{r}=\"\",\"\",C{r}/B{r})").font = NORMAL
    ws1.cell(r, 5, int(row["Census2011_Pop"])).font = BLUE_INPUT
    ws1.cell(r, 6, f"=E{r}/B{r}").font = NORMAL
    asm = row["Assembly_Seats"]
    if pd.notna(asm):
        ws1.cell(r, 7, int(asm)).font = BLUE_INPUT
        ws1.cell(r, 8, f"=E{r}/G{r}").font = NORMAL
        ws1.cell(r, 9, f"=G{r}/B{r}").font = NORMAL
    else:
        ws1.cell(r, 7, "—").font = NORMAL
        ws1.cell(r, 8, "—").font = NORMAL
        ws1.cell(r, 9, "—").font = NORMAL
    for c in range(1, len(headers)+1):
        ws1.cell(r, c).border = BORDER
    if c2 := ws1.cell(r, 3).value:
        ws1.cell(r, 3).number_format = "#,##0"
    ws1.cell(r, 4).number_format = "#,##0"
    ws1.cell(r, 5).number_format = "#,##0"
    ws1.cell(r, 6).number_format = "#,##0"

last_row = start_row + len(df) - 1

# Totals row
trow = last_row + 1
ws1.cell(trow, 1, "ALL-INDIA TOTAL / AVERAGE").font = BOLD
ws1.cell(trow, 2, f"=SUM(B{start_row}:B{last_row})").font = BOLD
ws1.cell(trow, 3, f"=SUM(C{start_row}:C{last_row})").font = BOLD
ws1.cell(trow, 3).number_format = "#,##0"
ws1.cell(trow, 4, f"=C{trow}/B{trow}").font = BOLD
ws1.cell(trow, 4).number_format = "#,##0"
ws1.cell(trow, 5, f"=SUM(E{start_row}:E{last_row})").font = BOLD
ws1.cell(trow, 5).number_format = "#,##0"
ws1.cell(trow, 6, f"=E{trow}/B{trow}").font = BOLD
ws1.cell(trow, 6).number_format = "#,##0"
ws1.cell(trow, 7, f"=SUM(G{start_row}:G{last_row})").font = BOLD
for c in range(1, len(headers)+1):
    ws1.cell(trow, c).border = BORDER
    ws1.cell(trow, c).fill = PatternFill("solid", start_color="D9E1F2")

# Highest / lowest flags
hr1 = trow + 2
ws1.cell(hr1, 1, "Highest avg electors/seat:").font = BOLD
ws1.cell(hr1, 2, f"=INDEX(A{start_row}:A{last_row},MATCH(MAX(D{start_row}:D{last_row}),D{start_row}:D{last_row},0))").font = NORMAL
ws1.cell(hr1, 3, f"=MAX(D{start_row}:D{last_row})").font = NORMAL
ws1.cell(hr1, 3).number_format = "#,##0"
ws1.cell(hr1+1, 1, "Lowest avg electors/seat:").font = BOLD
ws1.cell(hr1+1, 2, f"=INDEX(A{start_row}:A{last_row},MATCH(MIN(D{start_row}:D{last_row}),D{start_row}:D{last_row},0))").font = NORMAL
ws1.cell(hr1+1, 3, f"=MIN(D{start_row}:D{last_row})").font = NORMAL
ws1.cell(hr1+1, 3).number_format = "#,##0"
ws1.cell(hr1+2, 1, "National average (unweighted, mean of states):").font = BOLD
ws1.cell(hr1+2, 3, f"=AVERAGE(D{start_row}:D{last_row})").font = NORMAL
ws1.cell(hr1+2, 3).number_format = "#,##0"

note_r = hr1 + 4
ws1.cell(note_r, 1, "Note: Electors_2024 aggregated from ECI constituency-level data; ~0.5% variance vs official 96.88cr national figure (see README).").font = NOTE_FONT

col_widths = [32, 14, 16, 16, 20, 16, 13, 16, 16]
for i, w in enumerate(col_widths, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A3"

wb.save("election_analysis.xlsx")
print("Saved sheet 1 & 2. Rows:", start_row, "to", last_row, "| totals row:", trow)
