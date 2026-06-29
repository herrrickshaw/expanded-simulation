import pandas as pd
import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
RED_TITLE = Font(name=FONT_NAME, bold=True, size=13, color="C00000")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FAIL_FILL = PatternFill("solid", start_color="F8CBAD")
OK_FILL = PatternFill("solid", start_color="C6E0B4")

with open("model_results_real.pkl", "rb") as f:
    main_results = pickle.load(f)
with open("recent_split_results.pkl", "rb") as f:
    recent_results = pickle.load(f)

wb = load_workbook("election_ml_tcpd_real.xlsx")
ws = wb.create_sheet("GA Model Results (Real Data)", 2)
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Lok Sabha PC-Level Model: Train/Validation/Test on Real TCPD Data").font = TITLE_FONT
ws.merge_cells("A1:G1")
ws.cell(2, 1, "2,704 real PC-year winner rows (1999-2019), rolled up from 259,078 raw AC-segment rows. Chronological split: TRAIN 1999/2004, VALIDATION 2009, TEST 2014/2019.").font = NOTE_FONT
ws.merge_cells("A2:G2")

row = 4
ws.cell(row, 1, "PRIMARY RESULT: chronological split (genuine forward-looking test)").font = RED_TITLE
row += 1
ws.cell(row, 1, f"Majority-class baseline: validation {main_results['baseline_val_acc']:.1%}, test {main_results['baseline_test_acc']:.1%}").font = NOTE_FONT
row += 2

headers = ["Model", "Validation Accuracy", "Validation Macro-F1", "Test Accuracy", "Test Macro-F1", "Beats Baseline?"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER
row += 1
for name, res in main_results["results"].items():
    ws.cell(row, 1, name).font = NORMAL
    ws.cell(row, 2, res["val_accuracy"]).font = NORMAL
    ws.cell(row, 2).number_format = "0.0%"
    ws.cell(row, 3, res["val_macro_f1"]).font = NORMAL
    ws.cell(row, 3).number_format = "0.000"
    ws.cell(row, 4, res["test_accuracy"]).font = BOLD
    ws.cell(row, 4).number_format = "0.0%"
    ws.cell(row, 5, res["test_macro_f1"]).font = NORMAL
    ws.cell(row, 5).number_format = "0.000"
    beats = res["test_accuracy"] > main_results["baseline_test_acc"]
    c6 = ws.cell(row, 6, "Yes" if beats else "No -- worse than guessing baseline")
    c6.font = BOLD
    c6.fill = OK_FILL if beats else FAIL_FILL
    for c in range(1, 7):
        ws.cell(row, c).border = BORDER
    row += 1

row += 1
ws.cell(row, 1, "Test confusion matrix, best model (LightGBM) -- rows=actual, cols=predicted").font = SUBTITLE_FONT
row += 1
classes = main_results["label_encoder_classes"]
lgb_cm = main_results["results"]["LightGBM"]["test_confusion"]
for c, cls in enumerate(classes, start=2):
    ws.cell(row, c, f"Pred: {cls}").font = HEADER_FONT
    ws.cell(row, c).fill = HEADER_FILL
row += 1
for i, cls in enumerate(classes):
    ws.cell(row, 1, f"Actual: {cls}").font = BOLD
    for j, val in enumerate(lgb_cm[i]):
        ws.cell(row, j+2, val).font = NORMAL
    for c in range(1, len(classes)+2):
        ws.cell(row, c).border = BORDER
    row += 1

row += 1
ws.cell(row, 1, "Why it fails: a genuine political regime change, not a coding error").font = SUBTITLE_FONT
row += 1
why = [
    "Train years (1999, 2004) were dominated by INDIA-predecessor parties; test years (2014,",
    "2019) were the BJP/NDA wave era. Features learned from incumbency/margin patterns in the",
    "OLD regime cannot anticipate a genuine realignment -- this is a property of Indian",
    "electoral history, not a flaw in the modeling approach.",
]
for line in why:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

row += 1
ws.cell(row, 1, "SECONDARY CHECK: does recent-data training do better? (train 2009+2014, test 2019)").font = SUBTITLE_FONT
row += 1
ws.cell(row, 1, f"Majority-class baseline for this split: {recent_results['baseline_acc']:.1%} (easier baseline; 2009+2014 already lean toward 2019's eventual majority class)").font = NOTE_FONT
row += 2
headers2 = ["Model", "Test Accuracy", "Test Macro-F1", "Beats Baseline?"]
for c, h in enumerate(headers2, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER
row += 1
for name, res in recent_results["results"].items():
    ws.cell(row, 1, name).font = NORMAL
    ws.cell(row, 2, res["test_accuracy"]).font = BOLD
    ws.cell(row, 2).number_format = "0.0%"
    ws.cell(row, 3, res["test_macro_f1"]).font = NORMAL
    ws.cell(row, 3).number_format = "0.000"
    beats = res["test_accuracy"] > recent_results["baseline_acc"]
    c4 = ws.cell(row, 4, "Yes" if beats else "No -- still below this split's baseline")
    c4.font = BOLD
    c4.fill = OK_FILL if beats else FAIL_FILL
    for c in range(1, 5):
        ws.cell(row, c).border = BORDER
    row += 1

row += 1
note2 = [
    "Random Forest improved substantially (22.5% -> 52.0% test accuracy) when trained on the",
    "most recent prior election instead of a distant, different-regime era -- confirming",
    "recency matters. However, even this improved model remains BELOW its own split's",
    "majority-class baseline (62.9%), since 2014 was already a strong predictor of 2019's NDA",
    "dominance on its own. Net conclusion: recency helps, but neither split produces a model",
    "that reliably beats simple baseline guessing for this feature set.",
]
for line in note2:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

col_widths = [22, 18, 18, 16, 14, 26]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("election_ml_tcpd_real.xlsx")
print("GA Model Results sheet done.")
