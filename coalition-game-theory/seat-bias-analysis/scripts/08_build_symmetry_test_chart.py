import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
NORMAL = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

df = pd.read_csv("/home/claude/coalition_model/gan_scenario_results.csv")
wb = load_workbook("gan_symmetry_test.xlsx")

ws = wb.create_sheet("All Scenarios Chart Data")
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "P(Tracked Bloc Majority) vs Total Seats, All Five Scenarios").font = TITLE_FONT
ws.merge_cells("A1:G1")
ws.cell(2, 1, "Pivoted for charting: rows = seat-totals, columns = scenarios.").font = NOTE_FONT
ws.merge_cells("A2:G2")

pivot = df.pivot(index="total_seats", columns="scenario", values="pct_tracked_majority").reset_index()
scen_order = ["S0_NDA_Real","S1_INDIA_Mirror","S2_Tossup","S3_INDIA_Small","S4_NDA_Small"]
pivot = pivot[["total_seats"] + scen_order]

headers = ["Total Seats"] + scen_order
for c, h in enumerate(headers, start=1):
    cell = ws.cell(4, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

start_row = 5
for i, row in pivot.iterrows():
    r = start_row + i
    ws.cell(r, 1, int(row["total_seats"])).font = NORMAL
    for j, scen in enumerate(scen_order, start=2):
        c = ws.cell(r, j, float(row[scen]))
        c.number_format = "0.0%"
        c.font = NORMAL
    for c in range(1, len(headers)+1):
        ws.cell(r, c).border = BORDER
last_row = start_row + len(pivot) - 1

col_widths = [12] + [20]*len(scen_order)
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

chart = LineChart()
chart.title = "P(Tracked Bloc Majority) vs Total Seats -- All Five Scenarios"
chart.y_axis.title = "P(majority)"
chart.x_axis.title = "Total seats"
data_ref = Reference(ws, min_col=2, max_col=1+len(scen_order), min_row=4, max_row=last_row)
cats_ref = Reference(ws, min_col=1, min_row=start_row, max_row=last_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 12
chart.width = 26
ws.add_chart(chart, f"A{last_row+3}")

wb.save("gan_symmetry_test.xlsx")
print(f"Chart sheet done. rows {start_row}-{last_row}")
