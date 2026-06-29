from openpyxl import load_workbook
from openpyxl.styles import Font

FONT_NAME = "Arial"
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)

wb = load_workbook("coalition_model.xlsx")

# Reorder: Dashboard, Monte Carlo Simulation, README, Scenarios, MWC, Banzhaf
order = ["Dashboard", "Monte Carlo Simulation", "README", "Scenarios", "Minimal Winning Coalitions", "Banzhaf Power Index"]
wb._sheets = [wb[name] for name in order]

# Append Monte Carlo methodology notes to README
ws0 = wb["README"]
last_row = ws0.max_row + 2

notes = [
    ("Monte Carlo randomization (NEW)", BOLD),
    ("  In addition to the 6 hand-picked scenarios, the 'Monte Carlo Simulation' sheet runs 20,000 randomized", NORMAL),
    ("  seat distributions to answer: across plausible permutations, how often is the outcome actually hung?", NORMAL),
    ("  Method: each party/bloc's seat count is drawn uniformly from a historically grounded range (e.g. NDA", NORMAL),
    ("  220-360, based on its 293-353 span across 2014/2019/2024-plus-2026-realignment; YSRCP 0-25, reflecting", NORMAL),
    ("  its actual swing from 22 seats in 2019 to 0 in 2024). All draws are rescaled to sum to exactly 543.", NORMAL),
    ("  For every draw, the same exact game theory (minimal winning coalitions, Banzhaf power index) used in", NORMAL),
    ("  the hand-picked scenarios is recomputed. This is a structural sensitivity analysis, NOT a probabilistic", NORMAL),
    ("  forecast — it assumes a uniform (not poll-weighted) distribution within each party's plausible range,", NORMAL),
    ("  so the resulting percentages describe 'how much of the plausible-outcome SPACE is hung', not 'what is", NORMAL),
    ("  the probability of a hung parliament in 2029'. Random seed fixed (42) for reproducibility.", NORMAL),
]
r = last_row
for text, font in notes:
    ws0.cell(r, 1, text).font = font
    r += 1

wb.save("coalition_model.xlsx")
print("Reordered and README updated. New order:", wb.sheetnames)
