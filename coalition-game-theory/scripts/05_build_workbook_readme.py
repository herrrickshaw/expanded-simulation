import pandas as pd
import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

with open("scenario_results.pkl", "rb") as f:
    results = pickle.load(f)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, color="404040", italic=True)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
BLUE_INPUT = Font(name=FONT_NAME, size=10, color="0000FF")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HUNG_FILL = PatternFill("solid", start_color="F8CBAD")
MAJORITY_FILL = PatternFill("solid", start_color="C6E0B4")

wb = Workbook()

# ============================================================
# SHEET 1: README
# ============================================================
ws0 = wb.active
ws0.title = "README"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 112

r = 1
ws0.cell(r, 1, "Lok Sabha Coalition Game Theory Simulator").font = TITLE_FONT
r += 1
ws0.cell(r, 1, "Scenario analysis of majority formation using existing parties only — no new parties assumed.").font = SUBTITLE_FONT
r += 2

notes = [
    ("What this is", BOLD),
    ("  A structural simulation, NOT a forecast. It takes hypothetical seat distributions among parties", NORMAL),
    ("  that already exist today and asks: under coalition/cooperative game theory, which combinations of", NORMAL),
    ("  parties could form a majority (272 of 543 seats), and how much bargaining power does each party", NORMAL),
    ("  hold in that configuration? It does not predict vote shares, swing voters, or who would actually win.", NORMAL),
    ("", NORMAL),
    ("Game theory concepts used", BOLD),
    ("  • Majority threshold (quota) = 272 of 543 seats.", NORMAL),
    ("  • Minimal Winning Coalition (MWC): a set of parties whose combined seats >= 272, where removing ANY", NORMAL),
    ("    one member would drop the coalition below 272. These are the 'efficient' governing combinations —", NORMAL),
    ("    no party is carried for free; every member is necessary.", NORMAL),
    ("  • Banzhaf Power Index: for each party, the share of all possible coalitions in which that party is", NORMAL),
    ("    'critical' (its addition flips a losing coalition into a winning one). This captures bargaining", NORMAL),
    ("    leverage in a way raw seat count does not — a 4-seat party can have real power if it is the", NORMAL),
    ("    decisive vote in many possible combinations; a 200-seat party can have ZERO power if it can never", NORMAL),
    ("    be pivotal (e.g. if a single rival bloc already holds a majority alone).", NORMAL),
    ("", NORMAL),
    ("How parties are grouped", BOLD),
    ("  NDA and INDIA are treated as cohesive blocs (their real-world behavior to date) for the BASE CASE.", NORMAL),
    ("  Smaller unaligned/regional parties (YSRCP, AAP, AIMIM, SAD, etc.) are modeled individually since", NORMAL),
    ("  they are the parties most likely to be genuine kingmakers in a hung-parliament scenario.", NORMAL),
    ("  This mirrors how Indian coalition bargaining actually happens — alliance blocs negotiate as units;", NORMAL),
    ("  unaligned regional parties are courted individually.", NORMAL),
    ("", NORMAL),
    ("IMPORTANT correction from current actual standing (June 2026)", BOLD),
    ("  20 Trinamool Congress (TMC/AITC) MPs merged into the pre-existing Nationalist Citizens Party of", NORMAL),
    ("  India (NCPI, registered 2023) and joined NDA; 6 Shiv Sena (UBT) MPs migrated to Shiv Sena (also NDA).", NORMAL),
    ("  This pushed NDA's actual current strength to 318 seats (from 293 at the 2024 declaration) — already", NORMAL),
    ("  a standalone majority. Because NCPI is a pre-2026 registered party, this respects the 'no new", NORMAL),
    ("  parties' constraint. Scenario S0 reflects this current reality; S1 reflects the 2024 result AS", NORMAL),
    ("  DECLARED, before this realignment, for comparison.", NORMAL),
    ("", NORMAL),
    ("Scenarios modeled (see 'Scenarios' sheet for full detail)", BOLD),
    ("  S0: Actual current standing (NDA 318 — majority, no coalition math needed)", NORMAL),
    ("  S1: 2024 result as originally declared (NDA 293 — majority, but with allies, before TMC split)", NORMAL),
    ("  S2: NDA underperforms, INDIA also short — genuinely hung, regional parties needed by either side", NORMAL),
    ("  S3: Deep fragmentation — both blocs well under 272, several mid-sized regional parties pivotal", NORMAL),
    ("  S4: NDA narrowly short, one large regional revival (YSRCP-scale) becomes the deciding swing party", NORMAL),
    ("  S5: Role reversal — INDIA becomes largest bloc but still short of outright majority", NORMAL),
    ("", NORMAL),
    ("How to adjust scenarios yourself", BOLD),
    ("  Edit the BLUE seat-count cells on the 'Scenarios' sheet. Note: minimal winning coalitions and Banzhaf", NORMAL),
    ("  indices are NOT live Excel formulas (this combinatorial search is impractical in spreadsheet formulas);", NORMAL),
    ("  they were computed in Python for the scenarios shown. To test a new distribution, change the numbers,", NORMAL),
    ("  re-run the companion script, and the results will update. Ask for a re-run with your new numbers.", NORMAL),
    ("", NORMAL),
    ("Sources for current party seat data", BOLD),
    ("  Wikipedia 'List of members of the 18th Lok Sabha' (party-wise tally as of 28 March 2026) and", NORMAL),
    ("  Wikipedia 'National Democratic Alliance' (June 2026 TMC/NCPI merger and Shiv Sena migration update).", NORMAL),
    ("", NORMAL),
    ("Caveat", BOLD),
    ("  This is a methodological exercise built on illustrative seat numbers for scenarios S2-S5, not polling", NORMAL),
    ("  or projection data. Real coalition formation also depends on ideology, history, state-level rivalries,", NORMAL),
    ("  and bargaining beyond pure seat arithmetic — this captures the arithmetic layer only.", NORMAL),
]
for text, font in notes:
    ws0.cell(r, 1, text).font = font
    r += 1

wb.save("coalition_model.xlsx")
print("README done.")
