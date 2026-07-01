#!/usr/bin/env python3
"""
Build Electability Model Dashboard
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

REPO_ROOT = Path(__file__).parent.parent.parent
DATA_PROCESSED = REPO_ROOT / "data" / "processed"
OUTPUT_DIR = REPO_ROOT / "coalition-game-theory" / "output"

# Load analysis outputs
penalties_df = pd.read_csv(DATA_PROCESSED / "state_gender_penalties.csv")
const_potential_df = pd.read_csv(DATA_PROCESSED / "constituency_win_potential.csv")

# Sort
penalties_sorted = penalties_df.sort_values('Empirical_Gap_Pct')
const_high_potential = const_potential_df.nsmallest(100, 'Predicted_Gender_Gap')
const_best_representation = const_potential_df.nlargest(100, 'Predicted_Woman_Win_Prob')

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Dashboard"

# Title
row = 1
ws[f'A{row}'] = "Women's Electability Model"
ws[f'A{row}'].font = Font(size=14, bold=True)

row = 3
ws[f'A{row}'] = "Key Insight: Gender Penalty is Small Overall, Huge by State"
ws[f'A{row}'].font = Font(size=11, italic=True)

# Key stats
row = 5
stats_text = [
    ("Overall gender penalty (logistic coef):", "-0.0077 (women ~0.8% less likely to win)"),
    ("Incumbency boost", "+1.80 (incumbents 180% more likely)"),
    ("State with worst gender gap:", "Nagaland: -30.9% (no women elected ever)"),
    ("State with best gender gap:", "Assam: +86.7% (100% of women candidates won)"),
    ("", ""),
    ("Interpretation:", "Women face minimal structural disadvantage when they run."),
    ("", "The problem is supply (0.1% of candidates) not demand."),
    ("", "Reserved seats unlock latent candidate pools, don't overcome bias."),
]

for label, value in stats_text:
    ws[f'A{row}'] = label
    ws[f'B{row}'] = value
    if 'Interpretation' in label:
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
    row += 1

# State-level analysis
ws_state = wb.create_sheet("State Gender Penalties")
for r_idx, row_data in enumerate(dataframe_to_rows(penalties_sorted, index=False, header=True), 1):
    for c_idx, value in enumerate(row_data, 1):
        cell = ws_state.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        if isinstance(value, float):
            cell.number_format = '0.00'

ws_state.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws_state.column_dimensions[col].width = 18

# High potential constituencies
ws_potential = wb.create_sheet("High-Potential Constituencies")
for r_idx, row_data in enumerate(dataframe_to_rows(const_high_potential[['State', 'Constituency', 'Predicted_Woman_Win_Prob', 'Predicted_Gender_Gap', 'Empirical_Women_Win_Rate', 'Const_Type']], index=False, header=True), 1):
    for c_idx, value in enumerate(row_data, 1):
        cell = ws_potential.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        if isinstance(value, float):
            cell.number_format = '0.00'

ws_potential.column_dimensions['A'].width = 20
ws_potential.column_dimensions['B'].width = 30
for col in ['C', 'D', 'E', 'F']:
    ws_potential.column_dimensions[col].width = 18

# Best-representation constituencies
ws_repr = wb.create_sheet("Best Representation Potential")
for r_idx, row_data in enumerate(dataframe_to_rows(const_best_representation[['State', 'Constituency', 'Predicted_Woman_Win_Prob', 'Predicted_Gender_Gap', 'Empirical_Women_Win_Rate', 'Const_Type']], index=False, header=True), 1):
    for c_idx, value in enumerate(row_data, 1):
        cell = ws_repr.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        if isinstance(value, float):
            cell.number_format = '0.00'

ws_repr.column_dimensions['A'].width = 20
ws_repr.column_dimensions['B'].width = 30
for col in ['C', 'D', 'E', 'F']:
    ws_repr.column_dimensions[col].width = 18

# README
ws_readme = wb.create_sheet("README", 1)
readme = """Women's Electability Model (Phase 2a)

OVERVIEW
Logistic regression model (483k candidates, 56k winners) predicting win probability
from gender, incumbency, constituency type, turnout, and party.

SHEETS

1. Dashboard: Key statistics and interpretation
   - Gender penalty: -0.0077 coefficient (0.8% less likely to win)
   - Incumbency boost: +1.80 coefficient (180% more likely to win)
   - State-level variation: -30.9% (Nagaland) to +86.7% (Assam)

2. State Gender Penalties: All 34 states ranked by empirical gender gap
   - Women_Win_Rate_Pct: % of women candidates who won (empirical)
   - Men_Win_Rate_Pct: % of men candidates who won (empirical)
   - Empirical_Gap_Pct: women's rate minus men's rate (negative = disadvantaged)
   - Gender_Penalty_Coef: logistic regression coefficient (model-based)

3. High-Potential Constituencies: Top 100 with smallest predicted gender gap
   - Predicted_Gender_Gap: how much worse women are predicted to do (close to 0 = equal)
   - Predicted_Woman_Win_Prob: predicted win probability for woman candidate
   - These are candidates for reserved-seat placement

4. Best Representation Potential: Top 100 with highest predicted woman win rate
   - Predicted_Woman_Win_Prob: absolute win rate if woman candidate placed there
   - These constituencies would elect women without structural disadvantage

INTERPRETATION FOR DELIMITATION

1. Gender barrier is *state-specific*:
   - Some states (Nagaland, Mysore, Madras) have 0% women winners — supply problem
   - Others (Assam, Arunachal Pradesh) have 100% women winner rate — no barrier

2. Within-state variation:
   - High-potential constituencies show women can win with near-parity to men
   - These should be prioritized for reserved-seat placement

3. Incumbency is the dominant factor:
   - Being incumbent boosts win prob by 180% (far more than gender penalty of 0.8%)
   - Incumbent-displacement strategy crucial for rotation cycles

NEXT (Phase 2b)
Placement optimizer: assign ~33% reserved seats per state to maximize:
- Representation (elect women in high-potential constituencies)
- Coalition leverage (place women in swing states, make them pivotal)
- Geographic equity (spread across regions, caste, language)
"""

row = 1
for line in readme.split('\n'):
    ws_readme[f'A{row}'] = line
    ws_readme[f'A{row}'].alignment = Alignment(wrap_text=True)
    if line and line[0] not in ' \t':
        ws_readme.row_dimensions[row].height = 20
    row += 1

ws_readme.column_dimensions['A'].width = 100

# Save
output_file = OUTPUT_DIR / "Women_Electability_Model.xlsx"
wb.save(output_file)
print(f"✅ Saved: {output_file}")
print(f"\nSheets:")
print(f"  - Dashboard: Model overview")
print(f"  - State Gender Penalties: {len(penalties_df)} states")
print(f"  - High-Potential Constituencies: {len(const_high_potential)} constituencies")
print(f"  - Best Representation Potential: {len(const_best_representation)} constituencies")
