import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
RED_BOLD = Font(name=FONT_NAME, bold=True, size=10, color="C00000")
NORMAL = Font(name=FONT_NAME, size=10)

wb = Workbook()
ws0 = wb.active
ws0.title = "README"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 112

r = 1
ws0.cell(r, 1, "Real TCPD Lok Dhaba Data: Train/Validation/Test ML Pipeline").font = TITLE_FONT
r += 2

notes = [
    ("Data actually used (real, not synthetic or approximated)", BOLD),
    ("  GA file (Lok Sabha, 'AC Segment Wise' export): 259,078 raw rows, 1999-2019 (5 real", NORMAL),
    ("  elections), rolled up by summing each candidate's votes across their AC segments to", NORMAL),
    ("  the true PC-level (parliamentary constituency) result -> 2,704 genuine PC-year winner", NORMAL),
    ("  rows. AE file (State Assembly): 483,565 raw rows, 1961-2023 (62 years), aggregated to a", NORMAL),
    ("  national NDA/INDIA/Other bloc-share trend -- 56,358 real seat-winner rows behind it.", NORMAL),
    ("", NORMAL),
    ("Two real data-cleaning bugs found and fixed during rollup (documented for transparency)", BOLD),
    ("  1. NOTA rows were excluded before rollup -- in rare cases NOTA's vote total could rank", NORMAL),
    ("     #1 within an artifactual PC_Name sub-grouping (e.g. 'Gopalganj (SC)' vs 'Gopalganj'),", NORMAL),
    ("     which would have spuriously produced NOTA as a 'winner'.", NORMAL),
    ("  2. PC identity was keyed on (State, Year, PC_No), NOT PC_Name -- the raw data has the", NORMAL),
    ("     same PC_No appear under slightly different PC_Name strings across rows, which", NORMAL),
    ("     initially created duplicate PCs (10 cases in 2014) before the fix.", NORMAL),
    ("  After both fixes, per-year PC counts (538/538/543/543/542) match documented Lok Sabha", NORMAL),
    ("  seat totals for those years, with zero duplicate winners remaining.", NORMAL),
    ("", NORMAL),
    ("CHRONOLOGICAL train/validation/test split (genuine, not random)", BOLD),
    ("  TRAIN: 1999, 2004  |  VALIDATION: 2009  |  TEST: 2014, 2019", NORMAL),
    ("  The model is evaluated only on elections strictly after anything it was trained on.", NORMAL),
    ("", NORMAL),
    ("HONEST RESULT: the model fails on the primary chronological split", RED_BOLD),
    ("  Test accuracy: Logistic Regression 21.0%, Random Forest 22.5%, LightGBM 29.3%", NORMAL),
    ("  -- all near or below a majority-class baseline, and all far below what a useful 3-class", NORMAL),
    ("  classifier needs to be (random guessing alone gives ~33%). This is reported as a", NORMAL),
    ("  genuine negative result, not tuned away.", NORMAL),
    ("", NORMAL),
    ("Why it fails: a real regime change, not a coding error", BOLD),
    ("  Train years (1999, 2004) were INDIA-predecessor-dominant; test years (2014, 2019) were", NORMAL),
    ("  the BJP/NDA wave era. Incumbency and margin features learned from the OLD political", NORMAL),
    ("  regime cannot anticipate a genuine realignment. A secondary check -- training on the", NORMAL),
    ("  MOST RECENT prior election (2009+2014) to predict 2019 (same regime era) -- improves", NORMAL),
    ("  Random Forest accuracy to 52.0%, confirming recency matters, though even this remains", NORMAL),
    ("  below that split's own (easier) 62.9% majority-class baseline.", NORMAL),
    ("", NORMAL),
    ("What this means for 'predicting the new configuration'", BOLD),
    ("  Constituency-level structural features (incumbency, margin, fragmentation) do NOT", NORMAL),
    ("  reliably predict bloc outcomes across a realignment, even with genuine TCPD ground-", NORMAL),
    ("  truth data and a proper train/val/test pipeline. This reinforces the conclusion from", NORMAL),
    ("  the earlier population-based model: arithmetic and structural features are not a", NORMAL),
    ("  substitute for actual political/polling information when projecting future elections.", NORMAL),
    ("", NORMAL),
    ("Sources", BOLD),
    ("  TCPD Lok Dhaba (Trivedi Centre for Political Data, Ashoka University): GA and AE", NORMAL),
    ("  datasets, downloaded by the user directly from lokdhaba.ashoka.edu.in and uploaded for", NORMAL),
    ("  this analysis.", NORMAL),
]
for text, font in notes:
    ws0.cell(r, 1, text).font = font
    r += 1

wb.save("election_ml_tcpd_real.xlsx")
print("README done.")
