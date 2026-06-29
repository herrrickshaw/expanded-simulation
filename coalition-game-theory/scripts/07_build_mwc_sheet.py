import pandas as pd
import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("scenario_results.pkl", "rb") as f:
    results = pickle.load(f)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LIGHT_FILL = PatternFill("solid", start_color="EAF1F8")

wb = load_workbook("coalition_model.xlsx")
ws = wb.create_sheet("Minimal Winning Coalitions")
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Minimal Winning Coalitions by Scenario").font = TITLE_FONT
ws.merge_cells("A1:E1")
ws.cell(2, 1, "A Minimal Winning Coalition (MWC) is a set of parties reaching >=272 seats where every member is necessary — remove any one and it falls below majority.").font = NOTE_FONT
ws.merge_cells("A2:E2")

scenario_order = ["S0_Actual_Current","S1_2024_AsDeclared","S2_NDA_Falls_Short",
                  "S3_Deep_Hung_Parliament","S4_NDA_Needs_Swing","S5_INDIA_Largest_No_Majority"]
scenario_labels = {
    "S0_Actual_Current": "S0: Actual Current (June 2026)",
    "S1_2024_AsDeclared": "S1: 2024 Result As Declared",
    "S2_NDA_Falls_Short": "S2: NDA Falls Short, Hung",
    "S3_Deep_Hung_Parliament": "S3: Deep Fragmentation",
    "S4_NDA_Needs_Swing": "S4: NDA Needs One Swing Party",
    "S5_INDIA_Largest_No_Majority": "S5: INDIA Largest, No Majority",
}

row = 4
for s in scenario_order:
    ws.cell(row, 1, scenario_labels[s]).font = SUBTITLE_FONT
    ws.cell(row, 1).fill = LIGHT_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    headers = ["#", "Coalition", "Total Seats", "Margin Over 272", "# Parties Needed"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    row += 1

    mwcs = results[s]["minimal_winning_coalitions"]
    # Cap display: show all if <=20, else show smallest (most efficient) 20
    display_mwcs = mwcs if len(mwcs) <= 20 else sorted(mwcs, key=lambda x: (len(x["coalition"]), x["seats"]))[:20]

    for i, mwc in enumerate(display_mwcs, start=1):
        ws.cell(row, 1, i).font = NORMAL
        ws.cell(row, 2, " + ".join(mwc["coalition"])).font = NORMAL
        ws.cell(row, 3, mwc["seats"]).font = NORMAL
        ws.cell(row, 4, mwc["margin"]).font = NORMAL
        ws.cell(row, 5, len(mwc["coalition"])).font = NORMAL
        for c in range(1, 6):
            ws.cell(row, c).border = BORDER
        row += 1

    if len(mwcs) > 20:
        ws.cell(row, 1, f"... plus {len(mwcs)-20} more minimal winning coalitions (showing 20 smallest/most efficient)").font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    row += 2  # spacer between scenarios

col_widths = [6, 50, 14, 16, 18]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("coalition_model.xlsx")
print("Minimal Winning Coalitions sheet done.")
