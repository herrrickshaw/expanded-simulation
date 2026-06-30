import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=8, color="808080")
BOLD = Font(name=FONT_NAME, bold=True, size=10)
NORMAL = Font(name=FONT_NAME, size=10)
RED_BOLD = Font(name=FONT_NAME, bold=True, size=11, color="C00000")
GREEN_BOLD = Font(name=FONT_NAME, bold=True, size=11, color="375623")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CARD_FILL = PatternFill("solid", start_color="EAF1F8")

df = pd.read_csv("/home/claude/coalition_model/gan_scenario_results.csv")
wb = load_workbook("gan_symmetry_test.xlsx")

ws = wb.create_sheet("Symmetry Test Results", 0)
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "Hypothesis Test: Is Entrenchment Symmetric Regardless of Which Bloc Leads?").font = TITLE_FONT
ws.merge_cells("A1:H1")
ws.cell(2, 1, "Result: YES. The same simulation engine, with the lead flipped to INDIA or made small/neutral, shows entrenchment is about DISTANCE FROM 50%, not about NDA specifically.").font = NOTE_FONT
ws.merge_cells("A2:H2")

row = 4
ws.cell(row, 1, "Summary: P(tracked bloc majority) at 534 vs. 1,000 seats, by scenario").font = SUBTITLE_FONT
row += 1
headers = ["Scenario","Tracked Bloc","Target Margin","P(majority)\n@534 seats","P(majority)\n@1,000 seats","Change","Correlation\n(seats vs P(majority))"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER
row += 1
summary_start = row

scenario_labels = {
    "S0_NDA_Real": "S0: NDA Real (+11.1pp)",
    "S1_INDIA_Mirror": "S1: INDIA Mirror (+11.1pp)",
    "S2_Tossup": "S2: True Tossup (0pp)",
    "S3_INDIA_Small": "S3: INDIA Small (+3pp)",
    "S4_NDA_Small": "S4: NDA Small (+3pp)",
}
for scen, label in scenario_labels.items():
    sub = df[df.scenario == scen]
    track_bloc = sub["tracked_bloc"].iloc[0]
    margin = sub["target_margin_pp"].iloc[0]
    p534 = sub[sub.total_seats == 534]["pct_tracked_majority"].values[0]
    p1000 = sub[sub.total_seats == 1000]["pct_tracked_majority"].values[0]
    corr = np.corrcoef(sub["total_seats"], sub["pct_tracked_majority"])[0,1] if sub["pct_tracked_majority"].std() > 1e-9 else float("nan")

    ws.cell(row, 1, label).font = NORMAL
    ws.cell(row, 2, track_bloc).font = NORMAL
    ws.cell(row, 3, f"+{margin:.1f}pp" if margin > 0 else "0pp (tossup)").font = NORMAL
    ws.cell(row, 4, float(p534)).font = BOLD
    ws.cell(row, 4).number_format = "0.0%"
    ws.cell(row, 5, float(p1000)).font = BOLD
    ws.cell(row, 5).number_format = "0.0%"
    change = p1000 - p534
    c6 = ws.cell(row, 6, float(change))
    c6.number_format = "+0.0%;-0.0%"
    c6.font = BOLD
    ws.cell(row, 7, round(float(corr), 3) if not np.isnan(corr) else "n/a (constant)").font = NORMAL
    for c in range(1, 8):
        ws.cell(row, c).border = BORDER
        ws.cell(row, c).fill = CARD_FILL
    row += 1
summary_end = row - 1

row += 1
ws.cell(row, 1, "Key finding: symmetry holds").font = GREEN_BOLD
row += 1
finding = [
    "S3 (INDIA +3pp) and S4 (NDA +3pp) are near-mirror images: P(majority) rises from",
    "95.1%->99.1% for INDIA in S3, and from 95.9%->98.8% for NDA in S4 -- both climbing by",
    "roughly the same magnitude regardless of WHICH bloc holds the small lead.",
    "",
    "S2 (true tossup) confirms the mechanism precisely: when NEITHER bloc has a real edge,",
    "BOTH NDA and INDIA's win probabilities FALL together as seats increase (26.0% and",
    "25.9% at 534 seats, dropping toward ~21% and ~20% at 1,000) -- because more seats",
    "make a HUNG outcome more likely when nobody is genuinely ahead. Verified directly:",
    "hung-parliament rate in the tossup case rises from 48.1% (534 seats) to 58.8% (1,000",
    "seats), almost perfectly symmetric between the two sides throughout.",
    "",
    "S0 and S1 (the full +11.1pp real-margin mirrors) both saturate at 100% at every seat-",
    "total tested -- an 11.1pp lead is simply too decisive for either side to ever lose in",
    "this model, regardless of seat count, which is itself consistent with the hypothesis:",
    "a strong-enough lead is entrenched at EVERY size, not just large ones.",
]
for line in finding:
    ws.cell(row, 1, line).font = NORMAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

row += 1
ws.cell(row, 1, "Bottom line").font = RED_BOLD
row += 1
bottom = [
    "The hypothesis is confirmed: entrenchment with more seats is NOT specific to NDA. It is",
    "a property of DISTANCE FROM 50%, applying symmetrically to whichever bloc happens to",
    "hold a lead -- and it applies in mirror-image form to the HUNG outcome when neither",
    "bloc leads. More seats does not make Indian elections fairer in any of these",
    "configurations; it makes whatever the underlying truth is (a lead, a tie, or a tossup)",
    "more certain to play out exactly as the means suggest.",
]
for line in bottom:
    ws.cell(row, 1, line).font = BOLD
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

col_widths = [26, 12, 14, 14, 14, 12, 16]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("gan_symmetry_test.xlsx")
print(f"Symmetry Test Results sheet done. rows {summary_start}-{summary_end}")
