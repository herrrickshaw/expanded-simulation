from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pickle

with open("monte_carlo_summary.pkl", "rb") as f:
    summary = pickle.load(f)

FONT_NAME = "Arial"
BOLD = Font(name=FONT_NAME, bold=True, size=11)
NORMAL = Font(name=FONT_NAME, size=10)
BIG_NUM = Font(name=FONT_NAME, bold=True, size=18, color="1F4E78")
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HUNG_FILL = PatternFill("solid", start_color="F8CBAD")
MAJORITY_FILL = PatternFill("solid", start_color="C6E0B4")

wb = load_workbook("coalition_model.xlsx")
ws = wb["Dashboard"]

n_sims = summary["n_simulations"]
nda_maj = summary["nda_majority_count"]
india_maj = summary["india_majority_count"]
hung = summary["hung_count"]

# Insert rows after row 11 (end of scenario table) and before "Key Findings" (row 14)
ws.insert_rows(12, amount=6)

ws.cell(12, 1, f"Monte Carlo Check: Across {n_sims:,} Randomized Permutations").font = BOLD
ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=6)

labels = [
    ("NDA Outright Majority", nda_maj/n_sims, MAJORITY_FILL),
    ("INDIA Outright Majority", india_maj/n_sims, MAJORITY_FILL),
    ("Hung Parliament", hung/n_sims, HUNG_FILL),
]
col = 1
for label, pct, fill in labels:
    ws.cell(13, col, label).font = NORMAL
    ws.merge_cells(start_row=13, start_column=col, end_row=13, end_column=col+1)
    c = ws.cell(14, col, pct)
    c.font = BIG_NUM
    c.number_format = "0.0%"
    ws.merge_cells(start_row=14, start_column=col, end_row=15, end_column=col+1)
    for rr in (13,14,15):
        for cc in (col, col+1):
            ws.cell(rr, cc).fill = fill
            ws.cell(rr, cc).border = BORDER
    col += 2

ws.cell(17, 1, "See 'Monte Carlo Simulation' sheet for the full distribution and methodology.").font = Font(name=FONT_NAME, italic=True, size=8, color="808080")
ws.merge_cells(start_row=17, start_column=1, end_row=17, end_column=6)

wb.save("coalition_model.xlsx")
print("Dashboard updated with Monte Carlo summary.")
